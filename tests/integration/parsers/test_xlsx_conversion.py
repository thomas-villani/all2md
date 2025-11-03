"""Integration tests for XLSX to Markdown conversion."""

import pytest

try:
    import openpyxl
    from openpyxl.chart import BarChart, Reference

    HAS_OPENPYXL = True
except ImportError:
    openpyxl = None
    BarChart = Reference = None
    HAS_OPENPYXL = False

from all2md import to_ast, to_markdown
from all2md.ast.nodes import Document

pytestmark = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not available")


@pytest.mark.integration
def test_xlsx_to_markdown_basic_table(tmp_path):
    """Test basic XLSX to Markdown conversion with simple table."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Create a simple table
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "NYC"])
    ws.append(["Bob", 25, "LA"])
    ws.append(["Charlie", 35, "Chicago"])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    assert "Name" in result
    assert "Alice" in result
    assert "Bob" in result
    assert "Charlie" in result
    assert "NYC" in result


@pytest.mark.integration
def test_xlsx_to_markdown_single_column(tmp_path):
    """Test XLSX with single column conversion."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Item"])
    ws.append(["First"])
    ws.append(["Second"])
    ws.append(["Third"])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    assert "Item" in result
    assert "First" in result
    assert "Second" in result
    assert "Third" in result


@pytest.mark.integration
def test_xlsx_to_markdown_multiple_sheets(tmp_path):
    """Test XLSX with multiple sheets conversion."""
    wb = openpyxl.Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Data", "Value"])
    ws1.append(["A", 1])
    ws1.append(["B", 2])

    # Sheet 2
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Name", "Score"])
    ws2.append(["Alice", 95])
    ws2.append(["Bob", 88])

    # Sheet 3
    ws3 = wb.create_sheet("Sheet3")
    ws3.append(["Product", "Price"])
    ws3.append(["Widget", 10.99])
    ws3.append(["Gadget", 29.99])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    # Check for content from all sheets
    assert "Sheet1" in result or "Data" in result
    assert "Sheet2" in result or "Alice" in result
    assert "Sheet3" in result or "Widget" in result


@pytest.mark.integration
def test_xlsx_to_markdown_empty_cells(tmp_path):
    """Test XLSX with empty cells conversion."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Col1", "Col2", "Col3"])
    ws.append(["A", None, "C"])
    ws.append([None, "B", None])
    ws.append(["X", "Y", "Z"])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    assert "Col1" in result
    assert "Col2" in result
    assert "Col3" in result


@pytest.mark.integration
def test_xlsx_to_markdown_numeric_data(tmp_path):
    """Test XLSX with numeric data conversion."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Integer", "Float", "Percentage"])
    ws.append([100, 3.14, 0.85])
    ws.append([200, 2.71, 0.95])
    ws.append([300, 1.41, 0.75])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    assert "Integer" in result
    assert "Float" in result
    assert "100" in result
    assert "3.14" in result or "3" in result


@pytest.mark.integration
def test_xlsx_to_markdown_formulas(tmp_path):
    """Test XLSX with formulas conversion."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Value1", "Value2", "Sum"])
    ws.append([10, 20, "=A2+B2"])
    ws.append([30, 40, "=A3+B3"])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    assert "Value1" in result
    assert "Value2" in result
    assert "Sum" in result


@pytest.mark.integration
def test_xlsx_to_markdown_mixed_types(tmp_path):
    """Test XLSX with mixed data types."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["String", "Number", "Boolean", "Date"])
    ws.append(["Text", 123, True, "2025-01-01"])
    ws.append(["More text", 456, False, "2025-12-31"])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    assert "String" in result
    assert "Text" in result
    assert "123" in result


@pytest.mark.integration
def test_xlsx_to_markdown_large_table(tmp_path):
    """Test XLSX with large table conversion."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Create header
    ws.append(["Index", "Value", "Category"])

    # Add many rows
    for i in range(100):
        ws.append([i, i * 10, f"Category_{i % 5}"])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    assert "Index" in result
    assert "Value" in result
    assert "Category" in result


@pytest.mark.integration
def test_xlsx_to_markdown_unicode_content(tmp_path):
    """Test XLSX with Unicode characters."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Language", "Text"])
    ws.append(["Chinese", "\U00004e2d\U00006587"])
    ws.append(["Greek", "\U00000391\U000003b1"])
    ws.append(["Emoji", "\U0001f600 \U00002764"])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    assert "Language" in result
    assert "Text" in result


@pytest.mark.integration
def test_xlsx_to_markdown_empty_sheet(tmp_path):
    """Test XLSX with empty sheet conversion."""
    wb = openpyxl.Workbook()

    xlsx_file = tmp_path / "empty.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    # Should complete without error
    assert isinstance(result, str)


@pytest.mark.integration
def test_xlsx_to_markdown_special_characters(tmp_path):
    """Test XLSX with special characters in cells."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Special", "Characters"])
    ws.append(["Ampersand &", "Less than <"])
    ws.append(["Greater than >", 'Quotes ""'])
    ws.append(["Asterisks **", "Underscores __"])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    assert "Special" in result
    assert "Characters" in result


@pytest.mark.integration
def test_xlsx_to_markdown_with_chart(tmp_path):
    """Test XLSX with embedded chart."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Month", "Sales"])
    ws.append(["January", 100])
    ws.append(["February", 150])
    ws.append(["March", 200])

    # Add a chart
    chart = BarChart()
    chart.title = "Monthly Sales"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Sales"

    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    categories = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "D2")

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    # Check that data is present (chart may or may not be represented)
    assert "Month" in result
    assert "Sales" in result
    assert "January" in result


@pytest.mark.integration
def test_xlsx_to_markdown_merged_cells(tmp_path):
    """Test XLSX with merged cells."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Header1", "Header2", "Header3"])
    ws.append(["Data1", "Data2", "Data3"])

    # Merge cells
    ws.merge_cells("A1:B1")

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    # Should handle merged cells without error
    assert "Header" in result


@pytest.mark.integration
def test_xlsx_to_markdown_styled_cells(tmp_path):
    """Test XLSX with styled cells (bold, italic, colors)."""
    wb = openpyxl.Workbook()
    ws = wb.active

    from openpyxl.styles import Font

    ws.append(["Normal", "Bold", "Italic"])

    # Apply styles
    ws["B1"].font = Font(bold=True)
    ws["C1"].font = Font(italic=True)

    ws.append(["Regular text", "Bold text", "Italic text"])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    # Styles may or may not be preserved
    assert "Normal" in result
    assert "Bold" in result
    assert "Italic" in result


@pytest.mark.integration
def test_xlsx_to_ast_conversion(tmp_path):
    """Test XLSX to AST conversion pipeline."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Name", "Value"])
    ws.append(["Test", 123])

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    doc = to_ast(xlsx_file)

    # Verify AST structure
    assert isinstance(doc, Document)
    assert doc.children is not None

    # Verify content through markdown conversion
    result = to_markdown(xlsx_file)
    assert "Name" in result
    assert "Test" in result


@pytest.mark.integration
def test_xlsx_to_markdown_long_text_cells(tmp_path):
    """Test XLSX with cells containing long text."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Description", "Details"])
    ws.append(
        [
            "This is a very long description that spans many characters and words.",
            "This is also a long details field with lots of information.",
        ]
    )
    ws.append(
        [
            "Another long text entry with multiple sentences. It should be handled properly.",
            "More details here with additional information.",
        ]
    )

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    assert "Description" in result
    assert "Details" in result
    assert "very long description" in result


@pytest.mark.integration
def test_xlsx_to_markdown_hyperlinks(tmp_path):
    """Test XLSX with hyperlinks in cells."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Link Text", "URL"])
    ws["B2"].hyperlink = "https://example.com"
    ws["B2"] = "Example Site"

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    # Hyperlinks may or may not be preserved
    assert "Link Text" in result


@pytest.mark.integration
def test_xlsx_to_markdown_sparse_data(tmp_path):
    """Test XLSX with sparse data (non-contiguous cells)."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = "Top Left"
    ws["C3"] = "Middle"
    ws["E5"] = "Bottom Right"

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    # Should handle sparse data
    assert isinstance(result, str)


@pytest.mark.integration
def test_xlsx_to_markdown_column_widths(tmp_path):
    """Test XLSX with custom column widths."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Col1", "Col2", "Col3"])
    ws.append(["A", "B", "C"])

    # Set column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15

    xlsx_file = tmp_path / "test.xlsx"
    wb.save(str(xlsx_file))

    result = to_markdown(xlsx_file)

    # Column widths may not be preserved in markdown
    assert "Col1" in result
    assert "Col2" in result
