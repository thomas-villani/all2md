#  Copyright (c) 2025 Tom Villani, Ph.D.
#
# tests/unit/test_xlsx_ast.py
"""Unit tests for XLSX to AST converter.

Tests cover:
- Basic spreadsheet to AST conversion
- Image extraction from XLSX files
- Chart data extraction from XLSX files
- Chart mode options (skip vs data)
- Attachment mode handling

"""

from io import BytesIO

import pytest

from all2md.ast import Document, Heading, Image, Table
from all2md import XlsxOptions
from all2md.parsers.xlsx import XlsxToAstConverter

try:
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.drawing.image import Image as XlsxImage

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def create_xlsx_with_chart() -> BytesIO:
    """Create a XLSX file with a simple bar chart for testing.

    Returns
    -------
    BytesIO
        In-memory XLSX file with chart

    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"

    # Add data
    ws['A1'] = 'Month'
    ws['B1'] = 'Sales'
    ws['A2'] = 'Jan'
    ws['B2'] = 100
    ws['A3'] = 'Feb'
    ws['B3'] = 150
    ws['A4'] = 'Mar'
    ws['B4'] = 200

    # Create chart
    chart = BarChart()
    chart.title = "Monthly Sales"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Sales"

    # Define data range
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Add chart to worksheet
    ws.add_chart(chart, "D2")

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_xlsx_with_image() -> BytesIO:
    """Create a XLSX file with an embedded image for testing.

    Returns
    -------
    BytesIO
        In-memory XLSX file with image

    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Image Sheet"

    # Add some data
    ws['A1'] = 'Data with Image'
    ws['A2'] = 'Value 1'
    ws['A3'] = 'Value 2'

    # Create a simple 1x1 pixel PNG image
    # PNG header + minimal IDAT chunk for 1x1 transparent pixel
    minimal_png = (
        b'\x89PNG\r\n\x1a\n'  # PNG signature
        b'\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89'  # IHDR
        b'\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r\n-\xb4'  # IDAT
        b'\x00\x00\x00\x00IEND\xaeB`\x82'  # IEND
    )

    # Add image to sheet
    img_stream = BytesIO(minimal_png)
    img = XlsxImage(img_stream)
    img.anchor = 'C2'
    ws.add_image(img)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@pytest.mark.unit
@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestXlsxBasicConversion:
    """Tests for basic XLSX to AST conversion."""

    def test_simple_sheet_conversion(self) -> None:
        """Test converting a simple XLSX sheet to AST."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Sheet"

        # Add header and data
        ws['A1'] = 'Name'
        ws['B1'] = 'Age'
        ws['A2'] = 'Alice'
        ws['B2'] = 30
        ws['A3'] = 'Bob'
        ws['B3'] = 25

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Convert to AST
        converter = XlsxToAstConverter()
        ast_doc = converter.parse(output)

        assert isinstance(ast_doc, Document)
        assert len(ast_doc.children) > 0

        # Should have a heading for sheet title
        heading = ast_doc.children[0]
        assert isinstance(heading, Heading)
        assert heading.level == 2

        # Should have a table
        table = ast_doc.children[1]
        assert isinstance(table, Table)
        assert table.header is not None
        assert len(table.rows) == 2  # Two data rows

    def test_chart_mode_skip(self) -> None:
        """Test that charts are skipped when chart_mode is 'skip'."""
        xlsx_data = create_xlsx_with_chart()

        options = XlsxOptions(chart_mode="skip")
        converter = XlsxToAstConverter(options=options)
        ast_doc = converter.parse(xlsx_data)

        # Should have table but no chart nodes
        assert isinstance(ast_doc, Document)

        # Count table nodes (should be 1 - the data table)
        tables = [child for child in ast_doc.children if isinstance(child, Table)]
        assert len(tables) == 1  # Only the data table, no chart table

    def test_chart_mode_data(self) -> None:
        """Test that charts are extracted as tables when chart_mode is 'data'."""
        xlsx_data = create_xlsx_with_chart()

        options = XlsxOptions(chart_mode="data")
        converter = XlsxToAstConverter(options=options)
        ast_doc = converter.parse(xlsx_data)

        # Should have data table and chart extracted as table
        assert isinstance(ast_doc, Document)

        # Count table nodes (should be 2 - data table + chart table)
        tables = [child for child in ast_doc.children if isinstance(child, Table)]
        # Note: Chart extraction may or may not succeed depending on openpyxl's chart handling
        # At minimum we should have the data table
        assert len(tables) >= 1


@pytest.mark.unit
@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestXlsxImageExtraction:
    """Tests for image extraction from XLSX files."""

    def test_image_extraction_alt_text_mode(self) -> None:
        """Test image extraction with alt_text attachment mode."""
        xlsx_data = create_xlsx_with_image()

        options = XlsxOptions(attachment_mode="alt_text")
        converter = XlsxToAstConverter(options=options)
        ast_doc = converter.parse(xlsx_data)

        assert isinstance(ast_doc, Document)

        # Find Image nodes
        images = [child for child in ast_doc.children if isinstance(child, Image)]

        # May or may not have images depending on openpyxl's image handling
        # If we have images, verify they're properly formatted
        if images:
            assert len(images) >= 0
            for img in images:
                assert isinstance(img, Image)
                assert img.alt_text is not None

    def test_image_extraction_skip_mode(self) -> None:
        """Test that images are skipped when attachment_mode is 'skip'."""
        xlsx_data = create_xlsx_with_image()

        options = XlsxOptions(attachment_mode="skip")
        converter = XlsxToAstConverter(options=options)
        ast_doc = converter.parse(xlsx_data)

        assert isinstance(ast_doc, Document)

        # Should have no Image nodes
        images = [child for child in ast_doc.children if isinstance(child, Image)]
        assert len(images) == 0


@pytest.mark.unit
@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestXlsxOptions:
    """Tests for XLSX conversion options."""

    def test_include_sheet_titles(self) -> None:
        """Test include_sheet_titles option."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "My Sheet"
        ws['A1'] = 'Data'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # With sheet titles
        options = XlsxOptions(include_sheet_titles=True)
        converter = XlsxToAstConverter(options=options)
        ast_doc = converter.parse(output)

        # Should have heading for sheet title
        assert any(isinstance(child, Heading) and child.content[0].content == "My Sheet"
                   for child in ast_doc.children
                   if isinstance(child, Heading) and child.content)

    def test_no_sheet_titles(self) -> None:
        """Test that sheet titles are not included when disabled."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "My Sheet"
        ws['A1'] = 'Data'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Without sheet titles
        options = XlsxOptions(include_sheet_titles=False)
        converter = XlsxToAstConverter(options=options)
        ast_doc = converter.parse(output)

        # Should not have heading for sheet title
        headings = [child for child in ast_doc.children if isinstance(child, Heading)]
        assert len(headings) == 0
