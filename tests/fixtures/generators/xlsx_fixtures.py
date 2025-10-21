"""XLSX fixture generators using openpyxl for spreadsheet scenarios."""

from __future__ import annotations

from io import BytesIO

try:
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.drawing.image import Image as XlsxImage

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - optional dependency
    openpyxl = None  # type: ignore[assignment]
    BarChart = Reference = XlsxImage = None  # type: ignore[assignment]
    HAS_OPENPYXL = False

from utils import MINIMAL_PNG_BYTES


def create_xlsx_basic_table() -> bytes:
    """Create an XLSX workbook with a single sheet and simple table data."""
    _ensure_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(["Name", "Role", "Score"])
    ws.append(["Alice", "Engineer", 95])
    ws.append(["Bob", "Designer", 88])
    ws.append(["Carol", "Manager", 91])
    return _save_workbook_to_bytes(wb)


def create_xlsx_with_chart() -> bytes:
    """Create an XLSX workbook containing a bar chart for chart-mode tests."""
    _ensure_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"
    ws.append(["Month", "Revenue"])
    ws.append(["January", 100])
    ws.append(["February", 150])
    ws.append(["March", 200])

    chart = BarChart()
    chart.title = "Monthly Revenue"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Revenue"

    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    categories = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "D2")
    return _save_workbook_to_bytes(wb)


def create_xlsx_with_image(anchor: str = "C3") -> bytes:
    """Create an XLSX workbook embedding a minimal PNG image."""
    _ensure_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Images"
    ws.append(["Description", "Value"])
    ws.append(["Has image", 1])
    ws.append(["Another row", 2])

    image_stream = BytesIO(MINIMAL_PNG_BYTES)
    img = XlsxImage(image_stream)
    img.anchor = anchor
    ws.add_image(img)
    return _save_workbook_to_bytes(wb)


def create_xlsx_with_multiple_sheets() -> bytes:
    """Create an XLSX workbook featuring multiple sheets for traversal tests."""
    _ensure_openpyxl()
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Sheet", "Rows"])
    summary.append(["Summary", 2])
    summary.append(["Data", 4])

    data_sheet = wb.create_sheet("Data")
    data_sheet.append(["Product", "Units", "Price"])
    data_sheet.append(["Widget", 10, 9.99])
    data_sheet.append(["Gadget", 5, 19.99])
    data_sheet.append(["Doohickey", 3, 14.5])

    notes_sheet = wb.create_sheet("Notes")
    notes_sheet["A1"] = "Notes"
    notes_sheet["A2"] = "This sheet contains text notes for testing."

    return _save_workbook_to_bytes(wb)


def _save_workbook_to_bytes(workbook: "openpyxl.Workbook") -> bytes:
    """Serialize a workbook to bytes."""
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _ensure_openpyxl() -> None:
    """Raise an informative error if openpyxl is not installed."""
    if not HAS_OPENPYXL:  # pragma: no cover - executed only when missing
        raise ImportError(
            "openpyxl is required for XLSX fixture generation. Install dev dependencies to run these tests."
        )
