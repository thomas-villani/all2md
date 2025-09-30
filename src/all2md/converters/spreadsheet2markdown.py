#  Copyright (c) 2025 Tom Villani, Ph.D.
#
# src/all2md/converters/spreadsheet2markdown.py

"""Spreadsheet (XLSX/ODS/CSV/TSV) to Markdown conversion module.

This module converts:
- XLSX (Excel workbooks) via openpyxl
- ODS (OpenDocument Spreadsheets) via odfpy
- CSV and TSV via Python's csv module (with optional dialect detection)

Key Features
------------
- XLSX:
  - Per-sheet conversion to Markdown tables
  - Optional sheet selection by name list or regex
  - Hyperlink preservation ([text](url))
  - Best-effort handling for merged cells (blank out non-master cells)
    Note: Merged cells render as empty strings in non-master cell positions
    in the resulting Markdown table. Only the top-left master cell retains
    content. This is a limitation of Markdown's table format.
  - Optional formula rendering (values vs formulas) via data_only
  - Column alignment inferred from header cell horizontal alignment
- ODS:
  - Per-sheet conversion to Markdown tables
  - Optional sheet selection by name list or regex
  - Smart header detection using style analysis and content heuristics
  - Support for cell repetition attributes
  - Embedded image detection (framework ready for process_attachment)
- CSV/TSV:
  - Optional dialect detection (reads up to 4KB sample into memory)
  - Optional delimiter override (tsv uses tab)
  - First row used as header by default
  - Streaming parsing via csv.reader for memory efficiency with large files
- Enhanced header detection (all formats):
  - Manual mode (use has_header setting)
  - Auto mode (style-based heuristics for XLSX/ODS)
  - Numeric density mode (analyze text vs numeric content ratio)
- Row/column truncation with indicator
- Uses shared MarkdownOptions when available (escaping rules etc.)

Dependencies
------------
- openpyxl (only for XLSX)
- odfpy (only for ODS)
- CSV/TSV use standard library
"""

from __future__ import annotations

import csv
import io
import logging
import re
from pathlib import Path
from typing import IO, Any, Iterable, Optional, Union

from all2md.constants import TABLE_ALIGNMENT_MAPPING
from all2md.converter_metadata import ConverterMetadata
from all2md.exceptions import DependencyError, InputError, MarkdownConversionError
from all2md.options import MarkdownOptions, SpreadsheetOptions
from all2md.utils.inputs import escape_markdown_special, format_markdown_heading, validate_and_convert_input
from all2md.utils.metadata import (
    SPREADSHEET_FIELD_MAPPING,
    DocumentMetadata,
    map_properties_to_metadata,
    prepend_metadata_if_enabled,
)

logger = logging.getLogger(__name__)


class SpreadsheetConverterMetadata(ConverterMetadata):
    """Specialized metadata for spreadsheet converter with smart dependency checking."""

    def get_required_packages_for_content(self, content: Optional[bytes] = None) -> list[tuple[str, str]]:
        """Get required packages based on detected spreadsheet format.

        For XLSX files, openpyxl is required. For ODS files, odfpy is required.
        For CSV/TSV files, no additional packages are needed beyond the standard library.

        Parameters
        ----------
        content : bytes, optional
            File content to analyze for format detection

        Returns
        -------
        list[tuple[str, str]]
            Required packages for the detected format
        """
        if content is None:
            # If no content provided, assume worst case (both XLSX and ODS possible)
            return [("openpyxl", ""), ("odfpy", "")]

        # Check if it's a ZIP file (could be XLSX or ODS)
        if content.startswith(b'PK\x03\x04'):
            # Try to determine if it's ODS or XLSX
            try:
                import io
                import zipfile

                with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
                    file_list = zf.namelist()
                    # ODS files contain mimetype file with opendocument.spreadsheet
                    if 'mimetype' in file_list:
                        try:
                            mimetype = zf.read('mimetype').decode('utf-8', errors='ignore').strip()
                            if 'opendocument.spreadsheet' in mimetype:
                                return [("odfpy", "")]
                        except Exception:
                            pass
                    # XLSX files contain [Content_Types].xml and xl/ directory
                    if '[Content_Types].xml' in file_list or any(f.startswith('xl/') for f in file_list):
                        return [("openpyxl", "")]
                    # If we find META-INF/manifest.xml but no clear XLSX markers, likely ODS
                    if 'META-INF/manifest.xml' in file_list:
                        return [("odfpy", "")]
            except Exception:
                pass

            # Default to XLSX for ZIP files if we can't determine
            return [("openpyxl", "")]

        # Check if it's CSV/TSV via content detection
        if _detect_csv_tsv_content(content):
            return []  # CSV/TSV don't need additional packages

        # If we can't determine, assume XLSX for safety
        return [("openpyxl", "")]


def _sanitize_cell_text(text: Any, md_options: MarkdownOptions | None = None) -> str:
    """Convert any cell value to a safe Markdown table string."""
    if text is None:
        s = ""
    else:
        s = str(text)

    # Normalize whitespace/newlines inside cells
    s = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")

    # Escape markdown special characters if enabled
    if md_options and md_options.escape_special:
        s = escape_markdown_special(s)

    # Always escape the Markdown table pipe to avoid breaking the table
    s = s.replace("|", r"\|")
    return s


def _format_link_or_text(cell: Any, text: str, md_options: MarkdownOptions | None = None) -> str:
    """Render a cell as either a plain string or a Markdown link if hyperlink is present."""
    try:
        # openpyxl cell may have hyperlink attribute
        if hasattr(cell, "hyperlink") and cell.hyperlink and cell.hyperlink.target:
            disp = _sanitize_cell_text(text, md_options)
            url = cell.hyperlink.target
            return f"[{disp}]({url})"
    except Exception:
        pass

    return _sanitize_cell_text(text, md_options)


def _alignment_for_cell(cell: Any) -> str:
    """Map an openpyxl cell alignment to Markdown table alignment token."""
    try:
        align = getattr(cell, "alignment", None)
        if align and getattr(align, "horizontal", None):
            horiz = align.horizontal.lower()
            if horiz in TABLE_ALIGNMENT_MAPPING:
                return TABLE_ALIGNMENT_MAPPING[horiz]
    except Exception:
        pass

    # Default center (consistent with html2markdown default)
    return ":---:"


def _build_markdown_table(header: list[str], rows: list[list[str]], alignments: list[str]) -> str:
    """Assemble a Markdown table from header, rows, and column alignment tokens."""
    parts = []
    parts.append("| " + " | ".join(header) + " |")
    parts.append("|" + "|".join(alignments) + "|")
    for r in rows:
        parts.append("| " + " | ".join(r) + " |")
    return "\n".join(parts)


def _xlsx_iter_rows(sheet: Any, max_rows: int | None, max_cols: int | None) -> Iterable[list[Any]]:
    """Iterate rows of an openpyxl worksheet with optional truncation."""
    row_iter = sheet.iter_rows(values_only=False)  # need real cell objects
    for r_idx, row in enumerate(row_iter, start=1):
        if max_rows is not None and r_idx > max_rows:
            break
        out_row = []
        for c_idx, cell in enumerate(row, start=1):
            if max_cols is not None and c_idx > max_cols:
                break
            out_row.append(cell)
        yield out_row


def _map_merged_cells(sheet: Any) -> dict[str, str]:
    """Return a map of cell.coordinate -> master.coordinate for merged ranges."""
    merged_map: dict[str, str] = {}
    try:
        ranges = getattr(sheet, "merged_cells", None)
        if not ranges or not getattr(ranges, "ranges", None):
            return merged_map

        for mcr in ranges.ranges:
            min_r, min_c, max_r, max_c = mcr.min_row, mcr.min_col, mcr.max_row, mcr.max_col
            master = sheet.cell(row=min_r, column=min_c).coordinate
            for rr in range(min_r, max_r + 1):
                for cc in range(min_c, max_c + 1):
                    coord = sheet.cell(row=rr, column=cc).coordinate
                    merged_map[coord] = master
    except Exception as e:
        logger.debug(f"Unable to compute merged cell map: {e!r}")
    return merged_map


def extract_xlsx_metadata(workbook: Any) -> DocumentMetadata:
    """Extract metadata from XLSX workbook.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        The workbook object from openpyxl

    Returns
    -------
    DocumentMetadata
        Extracted metadata
    """
    if not hasattr(workbook, 'properties'):
        metadata = DocumentMetadata()
    else:
        props = workbook.properties
        # Use the utility function for standard metadata extraction
        metadata = map_properties_to_metadata(props, SPREADSHEET_FIELD_MAPPING)

        # Add XLSX-specific custom metadata
        custom_properties = ['lastModifiedBy', 'revision', 'version', 'company', 'manager']
        for prop_name in custom_properties:
            if hasattr(props, prop_name):
                value = getattr(props, prop_name)
                if value:
                    # Normalize property names for consistency
                    normalized_name = 'last_modified_by' if prop_name == 'lastModifiedBy' else prop_name
                    metadata.custom[normalized_name] = value

        # Application info as fallback creator
        if not metadata.creator and hasattr(props, 'application') and props.application:
            metadata.creator = props.application

    # Workbook-specific metadata
    if hasattr(workbook, 'sheetnames'):
        metadata.custom['sheet_count'] = len(workbook.sheetnames)
        metadata.custom['sheet_names'] = workbook.sheetnames

    return metadata


def _xlsx_to_markdown(workbook: Any, options: SpreadsheetOptions) -> str:
    """Convert an openpyxl workbook instance to Markdown across selected sheets."""
    md_options = options.markdown_options or MarkdownOptions()

    # Select sheets
    sheet_names: list[str] = list(workbook.sheetnames)
    if isinstance(options.sheets, list):
        sheet_names = [n for n in sheet_names if n in options.sheets]
    elif isinstance(options.sheets, str):
        pattern = re.compile(options.sheets)
        sheet_names = [n for n in sheet_names if pattern.search(n)]

    if not sheet_names:
        return ""

    sections: list[str] = []
    for sname in sheet_names:
        sheet = workbook[sname]

        if options.include_sheet_titles:
            use_hash = options.markdown_options.use_hash_headings if options.markdown_options else True
            sections.append(format_markdown_heading(sname, 2, use_hash))

        # Determine rows and merged cells
        merged_map = _map_merged_cells(sheet)

        raw_rows: list[list[Any]] = []
        for row in _xlsx_iter_rows(sheet, options.max_rows, options.max_cols):
            # Skip trailing empty rows (full row empty)
            if all((cell.value is None) for cell in row):
                raw_rows.append(row)  # keep row; user may prefer to see empty rows at top
            else:
                raw_rows.append(row)

        # Early-out for empty sheet
        if not raw_rows:
            continue

        # Convert to strings and handle merged cells
        str_rows: list[list[str]] = []
        for row in raw_rows:
            out: list[str] = []
            for cell in row:
                coord = getattr(cell, "coordinate", None)
                # If merged and not master, render as empty
                if coord and coord in merged_map and merged_map[coord] != coord:
                    out.append("")
                else:
                    out.append(_format_link_or_text(cell, cell.value, md_options))
            str_rows.append(out)

        # Trim completely empty trailing rows
        while str_rows and all(c == "" for c in str_rows[-1]):
            str_rows.pop()

        if not str_rows:
            continue

        header = str_rows[0]
        data = str_rows[1:] if len(str_rows) > 1 else []

        # Compute alignments from header cells
        alignments: list[str] = []
        try:
            first_row_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
            for c_idx, cell in enumerate(first_row_cells, start=1):
                if options.max_cols is not None and c_idx > options.max_cols:
                    break
                alignments.append(_alignment_for_cell(cell))
        except Exception:
            alignments = [":---:"] * len(header)

        table_md = _build_markdown_table(header, data, alignments)
        sections.append(table_md)

        # Truncation indicator
        truncated_rows = options.max_rows is not None and sheet.max_row > options.max_rows
        truncated_cols = options.max_cols is not None and sheet.max_column > options.max_cols
        if truncated_rows or truncated_cols:
            sections.append(f"\n*{options.truncation_indicator}*")

    # Separate sheets with a blank line or configured separator if desired
    return "\n\n".join([s.strip() for s in sections if s.strip()])


def xlsx_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[Any]],
        options: SpreadsheetOptions | None = None
) -> str:
    """Convert XLSX workbook to Markdown tables."""
    if options is None:
        options = SpreadsheetOptions()

    # Lazy import to have clean dependency errors
    try:
        import openpyxl
    except ImportError as e:
        raise DependencyError(
            converter_name="xlsx",
            missing_packages=[("openpyxl", "")]
        ) from e

    # Validate and open
    try:
        doc_input, input_type = validate_and_convert_input(
            input_data, supported_types=["path-like", "file-like", "bytes", "document objects"], require_binary=True
        )

        # If already an openpyxl workbook
        if input_type == "object" and getattr(doc_input, "__class__", None).__name__ == "Workbook":
            wb = doc_input
        else:
            wb = openpyxl.load_workbook(doc_input, data_only=options.render_formulas)

        # Extract metadata if requested
        metadata = None
        if options.extract_metadata:
            metadata = extract_xlsx_metadata(wb)

        result = _xlsx_to_markdown(wb, options)

        # Prepend metadata if enabled
        result = prepend_metadata_if_enabled(result, metadata, options.extract_metadata)

        return result

    except InputError:
        raise
    except Exception as e:
        raise MarkdownConversionError(
            f"Failed to process XLSX file: {e}", conversion_stage="document_opening", original_error=e
        ) from e


def _read_text_stream_for_csv(input_data: Any) -> io.StringIO:
    """Read binary or text input and return a StringIO for CSV parsing."""
    # If it's a StringIO already, return as-is
    if isinstance(input_data, io.StringIO):
        input_data.seek(0)
        return input_data

    # If it's a binary stream or bytes, decode
    content = None
    if hasattr(input_data, "read"):
        pos = None
        try:
            pos = input_data.tell()
        except Exception:
            pass

        raw = input_data.read()
        if isinstance(raw, bytes):
            # Try utf-8-sig first (to trim BOM if present)
            try:
                content = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                content = raw.decode("utf-8", errors="replace")
        else:
            content = str(raw)

        # Restore pointer if possible (not strictly required)
        try:
            if pos is not None:
                input_data.seek(pos)
        except Exception:
            pass
    else:
        # Fallback: treat as path-like or bytes already handled by validate_and_convert_input
        content = str(input_data)

    return io.StringIO(content or "")


def _csv_or_tsv_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[str]],
        options: SpreadsheetOptions | None = None,
        delimiter: str | None = None,
        force_delimiter: bool = False
) -> str:
    """Common CSV/TSV conversion using csv module."""
    if options is None:
        options = SpreadsheetOptions()
    md_options = options.markdown_options or MarkdownOptions()

    # Validate and load text
    try:
        doc_input, _ = validate_and_convert_input(input_data, supported_types=["path-like", "file-like", "bytes"])
        if isinstance(doc_input, (str, Path)):
            with open(doc_input, "rb") as f:
                text_stream = _read_text_stream_for_csv(f)
        else:
            text_stream = _read_text_stream_for_csv(doc_input)
    except InputError:
        raise
    except Exception as e:
        raise MarkdownConversionError(
            f"Failed to read CSV/TSV input: {e}", conversion_stage="input_processing", original_error=e
        ) from e

    # Sniff dialect if desired and not forcing a known delimiter
    text_stream.seek(0)
    sample = text_stream.read(4096)
    text_stream.seek(0)

    # Check if user provided a delimiter override via options
    if options.csv_delimiter:
        dialect = csv.excel()
        dialect.delimiter = options.csv_delimiter
    elif force_delimiter:
        dialect = csv.excel()
        dialect.delimiter = delimiter or "\t"
    elif options.detect_csv_dialect:
        try:
            sniffer = csv.Sniffer()
            detected = sniffer.sniff(sample, delimiters=[",", "\t", ";", "|", "\x1f"])
            dialect = detected
        except Exception:
            dialect = csv.excel()
            if delimiter:
                dialect.delimiter = delimiter
    else:
        dialect = csv.excel()
        if delimiter:
            dialect.delimiter = delimiter

    reader = csv.reader(text_stream, dialect=dialect)
    rows: list[list[str]] = []
    for r in reader:
        # Keep all rows; trimming happens below
        rows.append(r)

    # Drop leading/trailing fully empty rows
    def is_empty(row: list[str]) -> bool:
        return all((not (c or "").strip()) for c in row)

    while rows and is_empty(rows[0]):
        rows.pop(0)
    while rows and is_empty(rows[-1]):
        rows.pop()

    if not rows:
        return ""

    # Handle header based on has_header option
    if options.has_header:
        # Use first row as header
        header = rows[0]
        data_rows = rows[1:]
    else:
        # No header in data - generate generic headers based on first row column count
        if rows:
            num_cols = len(rows[0]) if rows else 0
            header = [f"Column {i+1}" for i in range(num_cols)]
            data_rows = rows  # All rows are data
        else:
            return ""

    # Truncate columns
    if options.max_cols is not None:
        header = header[: options.max_cols]
        data_rows = [r[: options.max_cols] for r in data_rows]

    # Truncate rows
    truncated = False
    if options.max_rows is not None and len(data_rows) > options.max_rows:
        data_rows = data_rows[: options.max_rows]
        truncated = True

    # Sanitize cells
    if options.has_header:
        # Only sanitize actual header text if it came from the file
        header = [_sanitize_cell_text(c, md_options).lstrip("\ufeff") for c in header]  # strip BOM in first header cell
    else:
        # Generic headers are already clean, just apply standard sanitization
        header = [_sanitize_cell_text(c, md_options) for c in header]
    data_rows = [[_sanitize_cell_text(c, md_options) for c in r] for r in data_rows]

    # Alignments default to center
    alignments = [":---:"] * len(header)

    table_md = _build_markdown_table(header, data_rows, alignments)

    if truncated:
        table_md += f"\n\n*{options.truncation_indicator}*"

    return table_md


def csv_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[str]],
        options: SpreadsheetOptions | None = None
) -> str:
    """Convert CSV to Markdown table."""
    return _csv_or_tsv_to_markdown(
        input_data=input_data,
        options=options,
        delimiter=",",
        force_delimiter=False  # allow dialect/sniffer unless overridden by options
    )


def tsv_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[str]],
        options: SpreadsheetOptions | None = None
) -> str:
    """Convert TSV to Markdown table."""
    # For TSV, default to tab delimiter and force it unless user enabled dialect detection explicitly
    force = True
    if options and options.detect_csv_dialect:
        force = False
    return _csv_or_tsv_to_markdown(
        input_data=input_data,
        options=options,
        delimiter="\t",
        force_delimiter=force
    )


def _detect_csv_tsv_content(content: bytes) -> bool:
    """Content-based detector for CSV/TSV formats.

    This function is used by the registry to detect CSV/TSV files
    based on content patterns when file extensions are not available.

    Parameters
    ----------
    content : bytes
        File content to analyze

    Returns
    -------
    bool
        True if content appears to be CSV or TSV
    """
    try:
        content_str = content.decode('utf-8', errors='ignore')
        non_empty_lines = [line.strip() for line in content_str.split('\n') if line.strip()]

        if len(non_empty_lines) >= 2:  # Need at least 2 lines (header + data)
            comma_count = sum(line.count(',') for line in non_empty_lines)
            tab_count = sum(line.count('\t') for line in non_empty_lines)

            # More relaxed CSV/TSV detection
            if comma_count >= len(non_empty_lines):  # At least one comma per line
                logger.debug(f"CSV pattern detected: {comma_count} commas in {len(non_empty_lines)} lines")
                return True
            elif tab_count >= len(non_empty_lines):  # At least one tab per line
                logger.debug(f"TSV pattern detected: {tab_count} tabs in {len(non_empty_lines)} lines")
                return True
    except UnicodeDecodeError:
        pass

    return False


def _detect_header_row_auto(rows: list[list[Any]], threshold: float = 0.7) -> int:
    """Auto-detect header row using heuristics.

    Parameters
    ----------
    rows : list[list[Any]]
        Rows of cell data
    threshold : float
        Minimum ratio of non-numeric cells to consider a row as header

    Returns
    -------
    int
        Row index that appears to be the header (0-based), or 0 if uncertain
    """
    if not rows or len(rows) < 2:
        return 0

    for row_idx, row in enumerate(rows[:3]):  # Check first 3 rows
        if not row:
            continue

        non_numeric_count = 0
        total_cells = 0

        for cell in row:
            if cell is not None and str(cell).strip():
                total_cells += 1
                # Check if cell value is non-numeric text
                cell_str = str(cell).strip()
                try:
                    float(cell_str)
                    # It's numeric
                except (ValueError, TypeError):
                    # It's text
                    non_numeric_count += 1

        if total_cells > 0:
            non_numeric_ratio = non_numeric_count / total_cells
            if non_numeric_ratio >= threshold:
                return row_idx

    # Default to first row
    return 0


def _ods_detect_header_row_style(sheet: Any, options: SpreadsheetOptions) -> int:
    """Detect header row using ODF style analysis.

    Parameters
    ----------
    sheet : odf.table.Table
        ODS table/sheet object
    options : SpreadsheetOptions
        Conversion options

    Returns
    -------
    int
        Row index that appears to be the header (0-based)
    """
    try:
        # Try to import odf modules
        from odf.table import TableCell, TableRow

        rows = list(sheet.getElementsByType(TableRow))
        if not rows:
            return 0

        # Check first few rows for bold formatting
        for row_idx, row in enumerate(rows[:3]):
            cells = list(row.getElementsByType(TableCell))
            if not cells:
                continue

            bold_count = 0
            total_text_cells = 0

            for cell in cells:
                text_content = str(cell).strip()
                if text_content:
                    total_text_cells += 1

                    # Check if cell has bold style
                    style_name = cell.getAttribute("stylename")
                    if style_name:
                        # This is a simplified check - in practice, we'd need to
                        # traverse the style hierarchy to check for bold formatting
                        # For now, we'll use a heuristic approach
                        if "bold" in style_name.lower() or "header" in style_name.lower():
                            bold_count += 1

            if total_text_cells > 0 and bold_count / total_text_cells >= 0.5:
                return row_idx

    except Exception as e:
        logger.debug(f"Error in style-based header detection: {e}")

    return 0


def _ods_extract_metadata(doc: Any) -> DocumentMetadata:
    """Extract metadata from ODS document.

    Parameters
    ----------
    doc : odf.opendocument.OpenDocument
        The ODS document object

    Returns
    -------
    DocumentMetadata
        Extracted metadata
    """
    metadata = DocumentMetadata()

    try:
        # Extract basic metadata
        meta = doc.meta
        if meta:
            # Get document info
            for child in meta.childNodes:
                if hasattr(child, 'qname'):
                    qname = child.qname
                    if qname and len(qname) >= 2:
                        tag_name = qname[1]  # Local name

                        if tag_name == "title" and child.firstChild:
                            metadata.title = str(child.firstChild.data)
                        elif tag_name == "creator" and child.firstChild:
                            metadata.creator = str(child.firstChild.data)
                        elif tag_name == "subject" and child.firstChild:
                            metadata.subject = str(child.firstChild.data)
                        elif tag_name == "description" and child.firstChild:
                            # Store description in custom metadata since DocumentMetadata doesn't have it
                            metadata.custom['description'] = str(child.firstChild.data)
                        elif tag_name == "creation-date" and child.firstChild:
                            metadata.creation_date = str(child.firstChild.data)
                        elif tag_name == "date" and child.firstChild:
                            metadata.modification_date = str(child.firstChild.data)

        # Add ODS-specific metadata
        body = doc.body
        if body:
            from odf.table import Table
            tables = list(body.getElementsByType(Table))
            metadata.custom['sheet_count'] = len(tables)
            metadata.custom['sheet_names'] = [
                table.getAttribute("name") or f"Sheet{i+1}" for i, table in enumerate(tables)
            ]

    except Exception as e:
        logger.debug(f"Error extracting ODS metadata: {e}")

    return metadata


def _ods_process_sheet(sheet: Any, sheet_name: str, options: SpreadsheetOptions) -> str:
    """Process a single ODS sheet into Markdown.

    Parameters
    ----------
    sheet : odf.table.Table
        ODS sheet/table object
    sheet_name : str
        Name of the sheet
    options : SpreadsheetOptions
        Conversion options

    Returns
    -------
    str
        Markdown representation of the sheet
    """
    try:
        from odf.draw import Frame
        from odf.table import TableCell, TableRow

        md_options = options.markdown_options or MarkdownOptions()
        sections = []

        # Add sheet title if requested
        if options.include_sheet_titles:
            use_hash = md_options.use_hash_headings if hasattr(md_options, 'use_hash_headings') else True
            sections.append(format_markdown_heading(sheet_name, 2, use_hash))

        # Extract all rows
        rows = list(sheet.getElementsByType(TableRow))
        if not rows:
            return ""

        # Convert to cell data
        raw_rows = []
        for row in rows:
            cells = list(row.getElementsByType(TableCell))
            row_data = []

            for cell in cells:
                # Get cell text content
                cell_text = ""
                for node in cell.childNodes:
                    if hasattr(node, 'data'):
                        cell_text += str(node.data)
                    elif hasattr(node, 'childNodes'):
                        for subnode in node.childNodes:
                            if hasattr(subnode, 'data'):
                                cell_text += str(subnode.data)

                # Handle cell repetition (number-columns-repeated attribute)
                repeat_count = 1
                try:
                    repeat_attr = cell.getAttribute("numbercolumnsrepeated")
                    if repeat_attr:
                        repeat_count = int(repeat_attr)
                except (ValueError, TypeError):
                    pass

                # Add cell data (with repetition)
                for _ in range(repeat_count):
                    row_data.append(cell_text.strip() if cell_text else "")

            raw_rows.append(row_data)

        # Remove empty trailing rows
        while raw_rows and all(not cell for cell in raw_rows[-1]):
            raw_rows.pop()

        if not raw_rows:
            return ""

        # Apply row limits
        if options.max_rows is not None:
            total_available = len(raw_rows) - 1  # Exclude potential header
            if total_available > options.max_rows:
                raw_rows = raw_rows[:options.max_rows + 1]  # +1 for header

        # Determine header row
        header_row_idx = 0
        if options.header_detection_mode == "auto":
            header_row_idx = _ods_detect_header_row_style(sheet, options)
        elif options.header_detection_mode == "numeric_density":
            header_row_idx = _detect_header_row_auto(raw_rows, options.auto_header_threshold)
        elif options.has_header:
            header_row_idx = 0
        else:
            # Generate synthetic header
            if raw_rows:
                max_cols = max(len(row) for row in raw_rows)
                if options.max_cols is not None:
                    max_cols = min(max_cols, options.max_cols)
                header = [f"Column {i+1}" for i in range(max_cols)]
                data_rows = raw_rows
            else:
                header = []
                data_rows = []

        if options.has_header or options.header_detection_mode in ("auto", "numeric_density"):
            if header_row_idx < len(raw_rows):
                header = raw_rows[header_row_idx]
                data_rows = raw_rows[header_row_idx + 1:]
            else:
                header = raw_rows[0] if raw_rows else []
                data_rows = raw_rows[1:] if len(raw_rows) > 1 else []

        # Apply column limits
        if options.max_cols is not None:
            header = header[:options.max_cols]
            data_rows = [row[:options.max_cols] for row in data_rows]

        # Sanitize all cell content
        header = [_sanitize_cell_text(cell, md_options) for cell in header]
        data_rows = [[_sanitize_cell_text(cell, md_options) for cell in row] for row in data_rows]

        # Ensure all rows have same number of columns
        if header:
            max_cols = len(header)
            data_rows = [row + [""] * (max_cols - len(row)) for row in data_rows]

        # Build table
        if header:
            alignments = [":---:"] * len(header)  # Default to center alignment
            table_md = _build_markdown_table(header, data_rows, alignments)
            sections.append(table_md)

        # Add truncation indicator if needed
        truncated = (options.max_rows is not None and len(rows) - 1 > options.max_rows) or \
                   (options.max_cols is not None and any(len(row) > options.max_cols for row in raw_rows))
        if truncated:
            sections.append(f"\n*{options.truncation_indicator}*")

        # Process embedded images/drawings
        try:
            frames = list(sheet.getElementsByType(Frame))
            for _frame in frames:
                # This would require more complex processing with process_attachment
                # For now, we'll just note their presence
                pass
        except Exception as e:
            logger.debug(f"Error processing ODS drawings: {e}")

        return "\n\n".join([s.strip() for s in sections if s.strip()])

    except Exception as e:
        logger.debug(f"Error processing ODS sheet '{sheet_name}': {e}")
        return ""


def ods_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[Any]],
        options: SpreadsheetOptions | None = None
) -> str:
    """Convert ODS spreadsheet to Markdown tables.

    Parameters
    ----------
    input_data : Union[str, Path, IO[bytes], IO[Any]]
        ODS file to convert
    options : SpreadsheetOptions | None
        Conversion options

    Returns
    -------
    str
        Markdown representation of the spreadsheet

    Raises
    ------
    DependencyError
        If odfpy is not installed
    MarkdownConversionError
        If conversion fails
    """
    if options is None:
        options = SpreadsheetOptions()

    # Lazy import to have clean dependency errors
    try:
        from odf import opendocument
        from odf.table import Table
    except ImportError as e:
        raise DependencyError(
            converter_name="ods",
            missing_packages=[("odfpy", "")]
        ) from e

    try:
        # Validate and open
        doc_input, input_type = validate_and_convert_input(
            input_data, supported_types=["path-like", "file-like", "bytes"], require_binary=True
        )

        # Load ODS document
        if input_type == "object" and hasattr(doc_input, 'body'):
            # Already an OpenDocument object
            doc = doc_input
        else:
            doc = opendocument.load(doc_input)

        # Extract metadata if requested
        metadata = None
        if options.extract_metadata:
            metadata = _ods_extract_metadata(doc)

        # Get all tables/sheets
        body = doc.body
        tables = list(body.getElementsByType(Table)) if body else []

        if not tables:
            result = ""
        else:
            # Filter sheets based on options
            sheet_names = [table.getAttribute("name") or f"Sheet{i+1}" for i, table in enumerate(tables)]
            selected_tables = []
            selected_names = []

            if isinstance(options.sheets, list):
                # Filter by exact names
                for i, name in enumerate(sheet_names):
                    if name in options.sheets:
                        selected_tables.append(tables[i])
                        selected_names.append(name)
            elif isinstance(options.sheets, str):
                # Filter by regex pattern
                import re
                pattern = re.compile(options.sheets)
                for i, name in enumerate(sheet_names):
                    if pattern.search(name):
                        selected_tables.append(tables[i])
                        selected_names.append(name)
            else:
                # Include all sheets
                selected_tables = tables
                selected_names = sheet_names

            # Process each selected sheet
            sections = []
            for table, name in zip(selected_tables, selected_names, strict=False):
                sheet_md = _ods_process_sheet(table, name, options)
                if sheet_md.strip():
                    sections.append(sheet_md)

            result = "\n\n".join(sections)

        # Prepend metadata if enabled
        result = prepend_metadata_if_enabled(result, metadata, options.extract_metadata)

        return result

    except InputError:
        raise
    except DependencyError:
        raise
    except Exception as e:
        raise MarkdownConversionError(
            f"Failed to process ODS file: {e}", conversion_stage="document_processing", original_error=e
        ) from e


def _detect_spreadsheet_format(input_data: Union[str, Path, IO[bytes], IO[str]]) -> str:
    """Detect spreadsheet format from input data.

    Returns
    -------
    str
        Format name: "xlsx", "ods", "csv", or "tsv"
    """
    # If it's a path, check extension first
    if isinstance(input_data, (str, Path)):
        path = Path(input_data)
        ext = path.suffix.lower()
        if ext == ".xlsx":
            return "xlsx"
        elif ext == ".ods":
            return "ods"
        elif ext == ".csv":
            return "csv"
        elif ext == ".tsv":
            return "tsv"

    # Check if file object has a name attribute
    elif hasattr(input_data, 'name'):
        filename = getattr(input_data, 'name', None)
        if filename:
            path = Path(filename)
            ext = path.suffix.lower()
            if ext == ".xlsx":
                return "xlsx"
            elif ext == ".ods":
                return "ods"
            elif ext == ".csv":
                return "csv"
            elif ext == ".tsv":
                return "tsv"

    # Try content-based detection
    try:
        # For file-like objects, read a sample
        if hasattr(input_data, 'read'):
            pos = getattr(input_data, 'tell', lambda: 0)()
            try:
                input_data.seek(0)
                sample = input_data.read(1024)
                input_data.seek(pos)  # Restore position
            except Exception:
                sample = b""
        elif isinstance(input_data, bytes):
            sample = input_data[:1024]
        else:
            # String path case - read first 1KB
            try:
                with open(input_data, 'rb') as f:
                    sample = f.read(1024)
            except Exception:
                sample = b""

        # Check for ZIP-based formats (XLSX or ODS)
        if sample.startswith(b'PK\x03\x04'):
            # Both XLSX and ODS are ZIP files, need to check content
            try:
                import io
                import zipfile

                # Create a file-like object from the input for ZIP inspection
                if hasattr(input_data, 'read'):
                    # For file-like objects, read the full content
                    pos = getattr(input_data, 'tell', lambda: 0)()
                    try:
                        input_data.seek(0)
                        zip_content = input_data.read()
                        input_data.seek(pos)  # Restore position
                    except Exception:
                        return "xlsx"  # Default to XLSX if we can't read
                elif isinstance(input_data, bytes):
                    zip_content = input_data
                else:
                    # String path case
                    try:
                        with open(input_data, 'rb') as f:
                            zip_content = f.read()
                    except Exception:
                        return "xlsx"  # Default to XLSX if we can't read

                # Check ZIP contents to distinguish XLSX from ODS
                try:
                    with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zf:
                        file_list = zf.namelist()
                        # ODS files contain META-INF/manifest.xml and mimetype
                        if 'mimetype' in file_list:
                            try:
                                mimetype = zf.read('mimetype').decode('utf-8', errors='ignore').strip()
                                if 'opendocument.spreadsheet' in mimetype:
                                    return "ods"
                            except Exception:
                                pass
                        # XLSX files contain [Content_Types].xml and xl/ directory
                        if '[Content_Types].xml' in file_list or any(f.startswith('xl/') for f in file_list):
                            return "xlsx"
                        # If we find META-INF/manifest.xml but no clear mimetype, likely ODS
                        if 'META-INF/manifest.xml' in file_list:
                            return "ods"
                except Exception:
                    pass
            except ImportError:
                pass

            # Default to XLSX for ZIP files if we can't determine the format
            return "xlsx"

        # Check for CSV/TSV patterns
        try:
            text_sample = sample.decode('utf-8', errors='ignore')
            lines = text_sample.split('\n')[:5]  # Check first 5 lines
            non_empty_lines = [line for line in lines if line.strip()]

            if len(non_empty_lines) >= 2:  # Need at least header + data
                comma_count = sum(line.count(',') for line in non_empty_lines)
                tab_count = sum(line.count('\t') for line in non_empty_lines)

                if tab_count >= len(non_empty_lines):  # At least one tab per line
                    return "tsv"
                elif comma_count >= len(non_empty_lines):  # At least one comma per line
                    return "csv"
        except UnicodeDecodeError:
            pass
    except Exception:
        pass

    # Default fallback to CSV
    return "csv"


def spreadsheet_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[str]],
        options: SpreadsheetOptions | None = None
) -> str:
    """Convert spreadsheet (XLSX/ODS/CSV/TSV) to Markdown table.

    This unified function detects the spreadsheet format and routes
    to the appropriate converter function.

    Parameters
    ----------
    input_data : Union[str, Path, IO[bytes], IO[str]]
        Input spreadsheet file or data
    options : SpreadsheetOptions | None, optional
        Conversion options

    Returns
    -------
    str
        Markdown table representation

    Raises
    ------
    DependencyError
        If required packages are not installed (openpyxl for XLSX, odfpy for ODS)
    InputError
        If input is invalid
    MarkdownConversionError
        If conversion fails
    """
    if options is None:
        options = SpreadsheetOptions()

    # Detect the actual format
    detected_format = _detect_spreadsheet_format(input_data)
    logger.debug(f"Detected spreadsheet format: {detected_format}")

    # Route to appropriate converter
    if detected_format == "xlsx":
        return xlsx_to_markdown(input_data, options)
    elif detected_format == "ods":
        return ods_to_markdown(input_data, options)
    elif detected_format == "csv":
        return csv_to_markdown(input_data, options)
    elif detected_format == "tsv":
        return tsv_to_markdown(input_data, options)
    else:
        # Fallback to CSV if detection fails
        logger.warning(f"Unknown format '{detected_format}', falling back to CSV")
        return csv_to_markdown(input_data, options)


CONVERTER_METADATA = SpreadsheetConverterMetadata(
    format_name="spreadsheet",
    extensions=[".xlsx", ".ods", ".csv", ".tsv"],
    mime_types=[
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # XLSX
        "application/vnd.oasis.opendocument.spreadsheet",  # ODS
        "text/csv",
        "application/csv",
        "text/tab-separated-values"  # TSV
    ],
    magic_bytes=[
        (b"PK\x03\x04", 0),  # ZIP signature for XLSX and ODS
    ],
    content_detector=_detect_csv_tsv_content,
    converter_module="all2md.converters.spreadsheet2markdown",
    converter_function="spreadsheet_to_markdown",
    required_packages=[("openpyxl", ""), ("odfpy", "")],  # Conditionally required based on format
    import_error_message="Spreadsheet conversion requires dependencies: 'openpyxl' for XLSX, "
                        "'odfpy' for ODS. Install with: pip install openpyxl odfpy",
    options_class="SpreadsheetOptions",
    description="Convert spreadsheets (XLSX, ODS, CSV, TSV) to Markdown tables",
    priority=5  # Reduced priority since ODS should be handled by ODF converter at priority 4 if both are available
)
