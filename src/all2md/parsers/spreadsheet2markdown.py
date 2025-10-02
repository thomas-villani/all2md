#  Copyright (c) 2025 Tom Villani, Ph.D.
#
# src/all2md/parsers/spreadsheet2markdown.py

"""Spreadsheet (XLSX/ODS/CSV/TSV) to Markdown conversion module.

This module converts:
- XLSX (Excel workbooks) via openpyxl
- ODS (OpenDocument Spreadsheets) via odfpy
- CSV and TSV via Python's csv module (with optional dialect detection)

Key Features
------------
- XLSX:
  - Per-sheet conversion to Markdown tables
  - Optional sheet selection by name list or regex
  - Hyperlink preservation ([text](url))
  - Best-effort handling for merged cells (blank out non-master cells)
    Note: Merged cells render as empty strings in non-master cell positions
    in the resulting Markdown table. Only the top-left master cell retains
    content. This is a limitation of Markdown's table format.
  - Optional formula rendering (values vs formulas) via data_only
  - Column alignment inferred from header cell horizontal alignment
- ODS:
  - Per-sheet conversion to Markdown tables
  - Optional sheet selection by name list or regex
  - Smart header detection using style analysis and content heuristics
  - Support for cell repetition attributes
  - Embedded image detection (framework ready for process_attachment)
- CSV/TSV:
  - Optional dialect detection (reads up to 4KB sample into memory)
  - Optional delimiter override (tsv uses tab)
  - First row used as header by default
  - Streaming parsing via csv.reader for memory efficiency with large files
- Enhanced header detection (all formats):
  - Manual mode (use has_header setting)
  - Auto mode (style-based heuristics for XLSX/ODS)
  - Numeric density mode (analyze text vs numeric content ratio)
- Row/column truncation with indicator
- Uses shared MarkdownOptions when available (escaping rules etc.)

Dependencies
------------
- openpyxl (only for XLSX)
- odfpy (only for ODS)
- CSV/TSV use standard library
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import IO, Any, Optional, Union

from all2md.converter_metadata import ConverterMetadata
from all2md.exceptions import DependencyError, InputError, MarkdownConversionError
from all2md.options import MarkdownOptions, SpreadsheetOptions
from all2md.utils.inputs import validate_and_convert_input
from all2md.utils.metadata import (
    SPREADSHEET_FIELD_MAPPING,
    DocumentMetadata,
    map_properties_to_metadata,
    prepend_metadata_if_enabled,
)

logger = logging.getLogger(__name__)


class SpreadsheetConverterMetadata(ConverterMetadata):
    """Specialized metadata for spreadsheet converter with smart dependency checking."""

    def get_required_packages_for_content(
        self,
        content: Optional[bytes] = None,
        input_data: Optional[Union[str, Path, IO[bytes], bytes]] = None
    ) -> list[tuple[str, str, str]]:
        """Get required packages based on detected spreadsheet format.

        For XLSX files, openpyxl is required. For ODS files, odfpy is required.
        For CSV/TSV files, no additional packages are needed beyond the standard library.

        Parameters
        ----------
        content : bytes, optional
            File content sample (may be partial) to analyze for format detection
        input_data : various types, optional
            Original input data (path, file object, or bytes) for accurate detection

        Returns
        -------
        list[tuple[str, str, str]]
            Required packages for the detected format as (install_name, import_name, version_spec) tuples
        """
        # If we have input_data, use it for accurate detection (can access full file)
        if input_data is not None:
            detected_format = _detect_spreadsheet_format(input_data)
            if detected_format == "xlsx":
                return [("openpyxl", "openpyxl", "")]
            elif detected_format == "ods":
                return [("odfpy", "odf", "")]
            elif detected_format in ("csv", "tsv"):
                return []  # CSV/TSV don't need additional packages

        # Fall back to content-based detection if no input_data
        if content is None:
            # If no content provided, assume worst case (both XLSX and ODS possible)
            return [("openpyxl", "openpyxl", ""), ("odfpy", "odf", "")]

        # Check if it's a ZIP file (could be XLSX or ODS)
        if content.startswith(b'PK\x03\x04'):
            # Check if content is large enough for ZIP operations (need more than just header)
            # ZIP files have central directory at the end, so need sufficient content
            # Minimum viable ZIP is about 22 bytes (for empty archive), but realistic files need more
            MIN_ZIP_SIZE = 100  # Minimum bytes needed for reliable ZIP inspection

            if len(content) >= MIN_ZIP_SIZE:
                # Try to determine if it's ODS or XLSX by opening ZIP
                try:
                    import zipfile

                    with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
                        file_list = zf.namelist()
                        # ODS files contain mimetype file with opendocument.spreadsheet
                        if 'mimetype' in file_list:
                            try:
                                mimetype = zf.read('mimetype').decode('utf-8', errors='ignore').strip()
                                if 'opendocument.spreadsheet' in mimetype:
                                    return [("odfpy", "odf", "")]
                            except Exception:
                                pass
                        # XLSX files contain [Content_Types].xml and xl/ directory
                        if '[Content_Types].xml' in file_list or any(f.startswith('xl/') for f in file_list):
                            return [("openpyxl", "openpyxl", "")]
                        # If we find META-INF/manifest.xml but no clear XLSX markers, likely ODS
                        if 'META-INF/manifest.xml' in file_list:
                            return [("odfpy", "odf", "")]
                except Exception:
                    # ZIP open failed (possibly truncated), be conservative
                    pass

            # Content is too short or ZIP open failed - be conservative
            # Return both packages since we can't reliably distinguish
            return [("openpyxl", "openpyxl", ""), ("odfpy", "odf", "")]

        # Check if it's CSV/TSV via content detection
        if _detect_csv_tsv_content(content):
            return []  # CSV/TSV don't need additional packages

        # If we can't determine, assume XLSX for safety
        return [("openpyxl", "openpyxl", "")]



def extract_xlsx_metadata(workbook: Any) -> DocumentMetadata:
    """Extract metadata from XLSX workbook.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        The workbook object from openpyxl

    Returns
    -------
    DocumentMetadata
        Extracted metadata
    """
    if not hasattr(workbook, 'properties'):
        metadata = DocumentMetadata()
    else:
        props = workbook.properties
        # Use the utility function for standard metadata extraction
        metadata = map_properties_to_metadata(props, SPREADSHEET_FIELD_MAPPING)

        # Add XLSX-specific custom metadata
        custom_properties = ['lastModifiedBy', 'revision', 'version', 'company', 'manager']
        for prop_name in custom_properties:
            if hasattr(props, prop_name):
                value = getattr(props, prop_name)
                if value:
                    # Normalize property names for consistency
                    normalized_name = 'last_modified_by' if prop_name == 'lastModifiedBy' else prop_name
                    metadata.custom[normalized_name] = value

        # Application info as fallback creator
        if not metadata.creator and hasattr(props, 'application') and props.application:
            metadata.creator = props.application

    # Workbook-specific metadata
    if hasattr(workbook, 'sheetnames'):
        metadata.custom['sheet_count'] = len(workbook.sheetnames)
        metadata.custom['sheet_names'] = workbook.sheetnames

    return metadata



def xlsx_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[Any]],
        options: SpreadsheetOptions | None = None
) -> str:
    """Convert XLSX workbook to Markdown tables."""
    if options is None:
        options = SpreadsheetOptions()

    # Lazy import to have clean dependency errors
    try:
        import openpyxl
    except ImportError as e:
        raise DependencyError(
            converter_name="xlsx",
            missing_packages=[("openpyxl", "")]
        ) from e

    # Validate and open
    try:
        doc_input, input_type = validate_and_convert_input(
            input_data, supported_types=["path-like", "file-like", "bytes", "document objects"], require_binary=True
        )

        # If already an openpyxl workbook
        doc_class = getattr(doc_input, "__class__", None)
        if input_type == "object" and doc_class is not None and doc_class.__name__ == "Workbook":
            wb = doc_input
        else:
            wb = openpyxl.load_workbook(doc_input, data_only=options.render_formulas)

        # Extract metadata if requested
        metadata = None
        if options.extract_metadata:
            metadata = extract_xlsx_metadata(wb)

        # Use AST-based conversion path
        from all2md.ast import MarkdownRenderer
        from all2md.parsers.spreadsheet import SpreadsheetToAstConverter

        # Convert to AST
        ast_converter = SpreadsheetToAstConverter(options=options)
        ast_document = ast_converter.xlsx_to_ast(wb)

        # Render AST to markdown
        md_opts = options.markdown_options if options.markdown_options else MarkdownOptions()
        renderer = MarkdownRenderer(md_opts)
        result = renderer.render(ast_document)

        # Prepend metadata if enabled
        result = prepend_metadata_if_enabled(result.strip(), metadata, options.extract_metadata)

        return result

    except InputError:
        raise
    except Exception as e:
        raise MarkdownConversionError(
            f"Failed to process XLSX file: {e}", conversion_stage="document_opening", original_error=e
        ) from e


def _read_text_stream_for_csv(input_data: Any) -> io.StringIO:
    """Read binary or text input and return a StringIO for CSV parsing."""
    # If it's a StringIO already, return as-is
    if isinstance(input_data, io.StringIO):
        input_data.seek(0)
        return input_data

    # If it's a binary stream or bytes, decode
    content = None
    if hasattr(input_data, "read"):
        pos = None
        try:
            pos = input_data.tell()
        except Exception:
            pass

        raw = input_data.read()
        if isinstance(raw, bytes):
            # Try utf-8-sig first (to trim BOM if present)
            try:
                content = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                content = raw.decode("utf-8", errors="replace")
        else:
            content = str(raw)

        # Restore pointer if possible (not strictly required)
        try:
            if pos is not None:
                input_data.seek(pos)
        except Exception:
            pass
    else:
        # Fallback: treat as path-like or bytes already handled by validate_and_convert_input
        content = str(input_data)

    return io.StringIO(content or "")



def csv_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[str]],
        options: SpreadsheetOptions | None = None
) -> str:
    """Convert CSV to Markdown table."""
    if options is None:
        options = SpreadsheetOptions()

    # Use AST-based conversion path
    from all2md.ast import MarkdownRenderer
    from all2md.parsers.spreadsheet import SpreadsheetToAstConverter

    ast_converter = SpreadsheetToAstConverter(options=options)
    ast_document = ast_converter.csv_or_tsv_to_ast(
        input_data=input_data,
        delimiter=",",
        force_delimiter=False
    )

    # Render AST to markdown
    md_opts = options.markdown_options if options.markdown_options else MarkdownOptions()
    renderer = MarkdownRenderer(md_opts)
    return renderer.render(ast_document)


def tsv_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[str]],
        options: SpreadsheetOptions | None = None
) -> str:
    """Convert TSV to Markdown table.

    By default, TSV files are parsed with tab (\\t) as the delimiter without
    dialect detection. Dialect detection only runs if options.detect_csv_dialect
    is explicitly set to True.

    Parameters
    ----------
    input_data : Union[str, Path, IO[bytes], IO[str]]
        TSV file to convert
    options : SpreadsheetOptions | None, default None
        Conversion options. Note: TSV forces tab delimiter unless
        detect_csv_dialect=True is set.

    Returns
    -------
    str
        Markdown representation of the TSV table

    Notes
    -----
    TSV delimiter behavior:
    - Default: Forces tab (\\t) delimiter
    - With detect_csv_dialect=True: Attempts automatic dialect detection
    - With csv_delimiter set: Uses specified delimiter (overrides default)

    Examples
    --------
    Default TSV conversion (forces tab delimiter):

        >>> tsv_to_markdown("data.tsv")

    With explicit dialect detection:

        >>> options = SpreadsheetOptions(detect_csv_dialect=True)
        >>> tsv_to_markdown("data.tsv", options)
    """
    if options is None:
        options = SpreadsheetOptions()

    # For TSV, default to tab delimiter and force it unless user enabled dialect detection explicitly
    force = True
    if options and options.detect_csv_dialect:
        force = False

    # Use AST-based conversion path
    from all2md.ast import MarkdownRenderer
    from all2md.parsers.spreadsheet import SpreadsheetToAstConverter

    ast_converter = SpreadsheetToAstConverter(options=options)
    ast_document = ast_converter.csv_or_tsv_to_ast(
        input_data=input_data,
        delimiter="\t",
        force_delimiter=force
    )

    # Render AST to markdown
    md_opts = options.markdown_options if options.markdown_options else MarkdownOptions()
    renderer = MarkdownRenderer(md_opts)
    return renderer.render(ast_document)


def _detect_csv_tsv_content(content: bytes) -> bool:
    """Content-based detector for CSV/TSV formats.

    This function is used by the registry to detect CSV/TSV files
    based on content patterns when file extensions are not available.

    Parameters
    ----------
    content : bytes
        File content to analyze

    Returns
    -------
    bool
        True if content appears to be CSV or TSV
    """
    try:
        content_str = content.decode('utf-8', errors='ignore')
        non_empty_lines = [line.strip() for line in content_str.split('\n') if line.strip()]

        if len(non_empty_lines) >= 2:  # Need at least 2 lines (header + data)
            comma_count = sum(line.count(',') for line in non_empty_lines)
            tab_count = sum(line.count('\t') for line in non_empty_lines)

            # More relaxed CSV/TSV detection
            if comma_count >= len(non_empty_lines):  # At least one comma per line
                logger.debug(f"CSV pattern detected: {comma_count} commas in {len(non_empty_lines)} lines")
                return True
            elif tab_count >= len(non_empty_lines):  # At least one tab per line
                logger.debug(f"TSV pattern detected: {tab_count} tabs in {len(non_empty_lines)} lines")
                return True
    except UnicodeDecodeError:
        pass

    return False



def _ods_extract_metadata(doc: Any) -> DocumentMetadata:
    """Extract metadata from ODS document.

    Parameters
    ----------
    doc : odf.opendocument.OpenDocument
        The ODS document object

    Returns
    -------
    DocumentMetadata
        Extracted metadata
    """
    metadata = DocumentMetadata()

    try:
        # Extract basic metadata
        meta = doc.meta
        if meta:
            # Get document info
            for child in meta.childNodes:
                if hasattr(child, 'qname'):
                    qname = child.qname
                    if qname and len(qname) >= 2:
                        tag_name = qname[1]  # Local name

                        if tag_name == "title" and child.firstChild:
                            metadata.title = str(child.firstChild.data)
                        elif tag_name == "creator" and child.firstChild:
                            metadata.creator = str(child.firstChild.data)
                        elif tag_name == "subject" and child.firstChild:
                            metadata.subject = str(child.firstChild.data)
                        elif tag_name == "description" and child.firstChild:
                            # Store description in custom metadata since DocumentMetadata doesn't have it
                            metadata.custom['description'] = str(child.firstChild.data)
                        elif tag_name == "creation-date" and child.firstChild:
                            metadata.creation_date = str(child.firstChild.data)
                        elif tag_name == "date" and child.firstChild:
                            metadata.modification_date = str(child.firstChild.data)

        # Add ODS-specific metadata
        body = doc.body
        if body:
            from odf.table import Table
            tables = list(body.getElementsByType(Table))
            metadata.custom['sheet_count'] = len(tables)
            metadata.custom['sheet_names'] = [
                table.getAttribute("name") or f"Sheet{i+1}" for i, table in enumerate(tables)
            ]

    except Exception as e:
        logger.debug(f"Error extracting ODS metadata: {e}")

    return metadata


def ods_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[Any]],
        options: SpreadsheetOptions | None = None
) -> str:
    """Convert ODS spreadsheet to Markdown tables.

    Parameters
    ----------
    input_data : Union[str, Path, IO[bytes], IO[Any]]
        ODS file to convert
    options : SpreadsheetOptions | None
        Conversion options

    Returns
    -------
    str
        Markdown representation of the spreadsheet

    Raises
    ------
    DependencyError
        If odfpy is not installed
    MarkdownConversionError
        If conversion fails
    """
    if options is None:
        options = SpreadsheetOptions()

    # Lazy import to have clean dependency errors
    try:
        from odf import opendocument
        from odf.table import Table
    except ImportError as e:
        raise DependencyError(
            converter_name="ods",
            missing_packages=[("odfpy", "")]
        ) from e

    try:
        # Validate and open
        doc_input, input_type = validate_and_convert_input(
            input_data, supported_types=["path-like", "file-like", "bytes"], require_binary=True
        )

        # Load ODS document
        if input_type == "object" and hasattr(doc_input, 'body'):
            # Already an OpenDocument object
            doc = doc_input
        else:
            doc = opendocument.load(doc_input)

        # Extract metadata if requested
        metadata = None
        if options.extract_metadata:
            metadata = _ods_extract_metadata(doc)

        # Use AST-based conversion path
        from all2md.ast import MarkdownRenderer
        from all2md.parsers.spreadsheet import SpreadsheetToAstConverter

        # Convert to AST
        ast_converter = SpreadsheetToAstConverter(options=options)
        ast_document = ast_converter.ods_to_ast(doc)

        # Render AST to markdown
        md_opts = options.markdown_options if options.markdown_options else MarkdownOptions()
        renderer = MarkdownRenderer(md_opts)
        result = renderer.render(ast_document)

        # Prepend metadata if enabled
        result = prepend_metadata_if_enabled(result.strip(), metadata, options.extract_metadata)

        return result

    except InputError:
        raise
    except DependencyError:
        raise
    except Exception as e:
        raise MarkdownConversionError(
            f"Failed to process ODS file: {e}", conversion_stage="document_processing", original_error=e
        ) from e


def _detect_spreadsheet_format(input_data: Union[str, Path, IO[bytes], IO[str], bytes]) -> str:
    """Detect spreadsheet format from input data.

    Returns
    -------
    str
        Format name: "xlsx", "ods", "csv", or "tsv"
    """
    # If it's a path, check extension first
    if isinstance(input_data, (str, Path)):
        path = Path(input_data)
        ext = path.suffix.lower()
        if ext == ".xlsx":
            return "xlsx"
        elif ext == ".ods":
            return "ods"
        elif ext == ".csv":
            return "csv"
        elif ext == ".tsv":
            return "tsv"

    # Check if file object has a name attribute
    elif hasattr(input_data, 'name'):
        filename = getattr(input_data, 'name', None)
        if filename:
            path = Path(filename)
            ext = path.suffix.lower()
            if ext == ".xlsx":
                return "xlsx"
            elif ext == ".ods":
                return "ods"
            elif ext == ".csv":
                return "csv"
            elif ext == ".tsv":
                return "tsv"

    # Try content-based detection
    try:
        # For file-like objects, read a sample
        sample: bytes = b""
        if hasattr(input_data, 'read') and not isinstance(input_data, (str, Path)):
            pos = getattr(input_data, 'tell', lambda: 0)()
            try:
                input_data.seek(0)
                sample_data = input_data.read(1024)
                input_data.seek(pos)
                # Ensure we have bytes
                if isinstance(sample_data, bytes):
                    sample = sample_data
                elif isinstance(sample_data, str):
                    sample = sample_data.encode('utf-8', errors='ignore')
            except Exception:
                pass
        elif isinstance(input_data, bytes):
            sample = input_data[:1024]
        elif isinstance(input_data, (str, Path)):
            # String or Path case - read first 1KB
            try:
                with open(str(input_data), 'rb') as f:
                    sample = f.read(1024)
            except Exception:
                pass

        # Check for ZIP-based formats (XLSX or ODS)
        if sample.startswith(b'PK\x03\x04'):
            # Both XLSX and ODS are ZIP files, need to check content
            try:
                import io
                import zipfile

                # Create a file-like object from the input for ZIP inspection
                zip_content: bytes = b""
                if hasattr(input_data, 'read') and not isinstance(input_data, (str, Path)):
                    # For file-like objects, read the full content
                    pos = getattr(input_data, 'tell', lambda: 0)()
                    try:
                        input_data.seek(0)
                        content_data = input_data.read()
                        input_data.seek(pos)
                        # Ensure we have bytes
                        if isinstance(content_data, bytes):
                            zip_content = content_data
                        elif isinstance(content_data, str):
                            zip_content = content_data.encode('utf-8', errors='ignore')
                    except Exception:
                        return "xlsx"  # Default to XLSX if we can't read
                elif isinstance(input_data, bytes):
                    zip_content = input_data
                elif isinstance(input_data, (str, Path)):
                    # String or Path case
                    try:
                        with open(str(input_data), 'rb') as f:
                            zip_content = f.read()
                    except Exception:
                        return "xlsx"  # Default to XLSX if we can't read

                # Check ZIP contents to distinguish XLSX from ODS
                try:
                    with zipfile.ZipFile(io.BytesIO(zip_content)) as zf:
                        file_list = zf.namelist()
                        # ODS files contain META-INF/manifest.xml and mimetype
                        if 'mimetype' in file_list:
                            try:
                                mimetype = zf.read('mimetype').decode('utf-8', errors='ignore').strip()
                                if 'opendocument.spreadsheet' in mimetype:
                                    return "ods"
                            except Exception:
                                pass
                        # XLSX files contain [Content_Types].xml and xl/ directory
                        if '[Content_Types].xml' in file_list or any(f.startswith('xl/') for f in file_list):
                            return "xlsx"
                        # If we find META-INF/manifest.xml but no clear mimetype, likely ODS
                        if 'META-INF/manifest.xml' in file_list:
                            return "ods"
                except Exception:
                    pass
            except ImportError:
                pass

            # Default to XLSX for ZIP files if we can't determine the format
            return "xlsx"

        # Check for CSV/TSV patterns
        try:
            text_sample = sample.decode('utf-8', errors='ignore')
            lines = text_sample.split('\n')[:5]  # Check first 5 lines
            non_empty_lines = [line for line in lines if line.strip()]

            if len(non_empty_lines) >= 2:  # Need at least header + data
                comma_count = sum(line.count(',') for line in non_empty_lines)
                tab_count = sum(line.count('\t') for line in non_empty_lines)

                if tab_count >= len(non_empty_lines):  # At least one tab per line
                    return "tsv"
                elif comma_count >= len(non_empty_lines):  # At least one comma per line
                    return "csv"
        except UnicodeDecodeError:
            pass
    except Exception:
        pass

    # Default fallback to CSV
    return "csv"


def spreadsheet_to_markdown(
        input_data: Union[str, Path, IO[bytes], IO[str]],
        options: SpreadsheetOptions | None = None
) -> str:
    """Convert spreadsheet (XLSX/ODS/CSV/TSV) to Markdown table.

    This unified function detects the spreadsheet format and routes
    to the appropriate converter function.

    Parameters
    ----------
    input_data : Union[str, Path, IO[bytes], IO[str]]
        Input spreadsheet file or data
    options : SpreadsheetOptions | None, optional
        Conversion options

    Returns
    -------
    str
        Markdown table representation

    Raises
    ------
    DependencyError
        If required packages are not installed (openpyxl for XLSX, odfpy for ODS)
    InputError
        If input is invalid
    MarkdownConversionError
        If conversion fails
    """
    if options is None:
        options = SpreadsheetOptions()

    # Detect the actual format
    detected_format = _detect_spreadsheet_format(input_data)
    logger.debug(f"Detected spreadsheet format: {detected_format}")

    # Route to appropriate converter
    if detected_format == "xlsx":
        return xlsx_to_markdown(input_data, options)
    elif detected_format == "ods":
        return ods_to_markdown(input_data, options)
    elif detected_format == "csv":
        return csv_to_markdown(input_data, options)
    elif detected_format == "tsv":
        return tsv_to_markdown(input_data, options)
    else:
        # Fallback to CSV if detection fails
        logger.warning(f"Unknown format '{detected_format}', falling back to CSV")
        return csv_to_markdown(input_data, options)


CONVERTER_METADATA = SpreadsheetConverterMetadata(
    format_name="spreadsheet",
    extensions=[".xlsx", ".ods", ".csv", ".tsv"],
    mime_types=[
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # XLSX
        "application/vnd.oasis.opendocument.spreadsheet",  # ODS
        "text/csv",
        "application/csv",
        "text/tab-separated-values"  # TSV
    ],
    magic_bytes=[
        (b"PK\x03\x04", 0),  # ZIP signature for XLSX and ODS
    ],
    content_detector=_detect_csv_tsv_content,
    converter_module="all2md.parsers.spreadsheet2markdown",
    converter_function="spreadsheet_to_markdown",
    required_packages=[("openpyxl", "openpyxl", ""), ("odfpy", "odf", "")],  # Conditionally required based on format
    import_error_message="Spreadsheet conversion requires dependencies: 'openpyxl' for XLSX, "
                        "'odfpy' for ODS. Install with: pip install openpyxl odfpy",
    options_class="SpreadsheetOptions",
    description="Convert spreadsheets (XLSX, ODS, CSV, TSV) to Markdown tables",
    priority=5  # Reduced priority since ODS should be handled by ODF converter at priority 4 if both are available
)
