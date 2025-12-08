#  Copyright (c) 2025 Tom Villani, Ph.D.
#
# src/all2md/parsers/xlsx.py
"""XLSX to AST converter.

This module provides conversion from Excel XLSX files to AST representation.
It replaces the combined spreadsheet parser with a focused Excel parser.

"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import IO, Any, Iterable, Optional, Union, cast

from all2md.ast import (
    Alignment,
    Document,
    Heading,
    HTMLInline,
    Image,
    Node,
    Paragraph,
    Table,
    Text,
)
from all2md.constants import DEPS_XLSX
from all2md.converter_metadata import ConverterMetadata
from all2md.exceptions import MalformedFileError
from all2md.options.xlsx import XlsxOptions
from all2md.parsers.base import BaseParser
from all2md.progress import ProgressCallback
from all2md.utils.attachments import create_attachment_sequencer, process_attachment
from all2md.utils.chart_helpers import build_chart_table
from all2md.utils.decorators import requires_dependencies
from all2md.utils.inputs import validate_and_convert_input
from all2md.utils.metadata import DocumentMetadata
from all2md.utils.parser_helpers import attachment_result_to_image_node
from all2md.utils.spreadsheet import (
    build_table_ast,
    sanitize_cell_text,
    transform_header_case,
    trim_columns,
    trim_rows,
)

logger = logging.getLogger(__name__)


def _format_link_or_text(cell: Any, text: str, preserve_newlines: bool = False) -> str:
    """Get cell text, checking for hyperlinks.

    Parameters
    ----------
    cell : Any
        Cell object (openpyxl)
    text : str
        Cell text value
    preserve_newlines : bool
        Whether to preserve newlines in cells

    Returns
    -------
    str
        Cell text with hyperlinks formatted as markdown links if present

    """
    # Check if cell has a hyperlink
    if hasattr(cell, "hyperlink") and cell.hyperlink:
        # Extract the URL from the hyperlink
        url = None
        if hasattr(cell.hyperlink, "target") and cell.hyperlink.target:
            url = cell.hyperlink.target
        elif hasattr(cell.hyperlink, "location") and cell.hyperlink.location:
            # Internal link (cell reference)
            url = f"#{cell.hyperlink.location}"

        if url:
            # Format as markdown link: [text](url)
            sanitized_text = sanitize_cell_text(text, preserve_newlines)
            return f"[{sanitized_text}]({url})"

    # No hyperlink, return sanitized text
    return sanitize_cell_text(text, preserve_newlines)


def _alignment_for_cell(cell: Any) -> Alignment:
    """Map an openpyxl cell alignment to alignment string.

    Parameters
    ----------
    cell : Any
        Cell object

    Returns
    -------
    Alignment
        Alignment: 'left', 'center', 'right', or 'center' (default)

    """
    try:
        align = getattr(cell, "alignment", None)
        if align and getattr(align, "horizontal", None):
            horiz = align.horizontal.lower()
            # Map Excel alignment to our alignment values
            alignment_map: dict[str, Alignment] = {
                "left": "left",
                "center": "center",
                "centre": "center",
                "right": "right",
                "general": "left",
                "justify": "left",
            }
            return alignment_map.get(horiz, "center")
    except Exception:
        pass

    return "center"


def _xlsx_iter_rows(sheet: Any, max_rows: int | None, max_cols: int | None) -> Iterable[list[Any]]:
    """Iterate rows of an openpyxl worksheet with optional truncation.

    Parameters
    ----------
    sheet : Any
        Openpyxl worksheet
    max_rows : int | None
        Maximum rows to read
    max_cols : int | None
        Maximum columns to read

    Yields
    ------
    list[Any]
        Row cells

    """
    row_iter = sheet.iter_rows(values_only=False)
    for r_idx, row in enumerate(row_iter, start=1):
        if max_rows is not None and r_idx > max_rows:
            break
        out_row = []
        for c_idx, cell in enumerate(row, start=1):
            if max_cols is not None and c_idx > max_cols:
                break
            out_row.append(cell)
        yield out_row


def _map_merged_cells(sheet: Any) -> dict[str, str]:
    """Return a map of cell.coordinate -> master.coordinate for merged ranges.

    Parameters
    ----------
    sheet : Any
        Openpyxl worksheet

    Returns
    -------
    dict[str, str]
        Mapping of cell coordinates to master cell coordinates

    """
    merged_map: dict[str, str] = {}
    try:
        ranges = getattr(sheet, "merged_cells", None)
        if not ranges or not getattr(ranges, "ranges", None):
            return merged_map

        for mcr in ranges.ranges:
            min_r, min_c, max_r, max_c = mcr.min_row, mcr.min_col, mcr.max_row, mcr.max_col
            master = sheet.cell(row=min_r, column=min_c).coordinate
            for rr in range(min_r, max_r + 1):
                for cc in range(min_c, max_c + 1):
                    coord = sheet.cell(row=rr, column=cc).coordinate
                    merged_map[coord] = master
    except Exception as e:
        logger.debug(f"Unable to compute merged cell map: {e!r}")
    return merged_map


def _get_merged_cell_spans(sheet: Any) -> dict[str, tuple[int, int]]:
    """Return a map of master cell coordinates to (colspan, rowspan) tuples.

    Parameters
    ----------
    sheet : Any
        Openpyxl worksheet

    Returns
    -------
    dict[str, tuple[int, int]]
        Mapping of master cell coordinates to (colspan, rowspan) tuples

    """
    span_map: dict[str, tuple[int, int]] = {}
    try:
        ranges = getattr(sheet, "merged_cells", None)
        if not ranges or not getattr(ranges, "ranges", None):
            return span_map

        for mcr in ranges.ranges:
            min_r, min_c, max_r, max_c = mcr.min_row, mcr.min_col, mcr.max_row, mcr.max_col
            master = sheet.cell(row=min_r, column=min_c).coordinate
            colspan = max_c - min_c + 1
            rowspan = max_r - min_r + 1
            span_map[master] = (colspan, rowspan)
    except Exception as e:
        logger.debug(f"Unable to compute merged cell spans: {e!r}")
    return span_map


def _extract_sheet_images(
    sheet: Any, base_filename: str, attachment_sequencer: Any, options: Any
) -> tuple[list[Image], dict[str, str]]:
    """Extract images from an XLSX sheet and convert to Image AST nodes.

    Parameters
    ----------
    sheet : Any
        Openpyxl worksheet
    base_filename : str
        Base filename for generating image filenames
    attachment_sequencer : callable
        Sequencer for generating unique attachment filenames
    options : XlsxOptions
        Conversion options

    Returns
    -------
    tuple[list[Image], dict[str, str]]
        Tuple of (list of Image AST nodes, footnotes dict)

    """
    images: list[Image] = []
    collected_footnotes: dict[str, str] = {}

    try:
        # NOTE: Using private API (sheet._images) because openpyxl does not provide
        # a public API for reading existing images from worksheets. This is the
        # standard approach per openpyxl documentation and community practice.
        # Compatible with openpyxl>=3.1.5. May need updates if openpyxl changes internals.
        if not hasattr(sheet, "_images") or not sheet._images:
            return images, collected_footnotes

        for img in sheet._images:
            try:
                # Get image data using private API methods (no public alternative exists)
                # Primary method: img._data() - documented approach for openpyxl>=3.1
                image_bytes = None
                if hasattr(img, "_data") and callable(img._data):
                    try:
                        image_bytes = img._data()
                    except Exception as e:
                        logger.debug(f"Failed to call img._data(): {e!r}")

                # Fallback method: access via relationships
                if not image_bytes and hasattr(img, "ref"):
                    image_part = img.ref
                    if hasattr(image_part, "blob"):
                        image_bytes = image_part.blob

                if not image_bytes:
                    logger.debug("Could not extract image data from XLSX sheet")
                    continue

                # Determine file extension
                extension = "png"  # default
                if hasattr(img, "format"):
                    extension = img.format.lower()
                elif hasattr(img, "ref") and hasattr(img.ref, "content_type"):
                    content_type = img.ref.content_type.lower()
                    if "jpeg" in content_type or "jpg" in content_type:
                        extension = "jpg"
                    elif "png" in content_type:
                        extension = "png"
                    elif "gif" in content_type:
                        extension = "gif"

                # Generate filename
                image_filename, _ = attachment_sequencer(
                    base_stem=base_filename, format_type="general", extension=extension, attachment_type="img"
                )

                # Get alt text
                alt_text = getattr(img, "name", "") or getattr(img, "title", "") or "image"

                # Process attachment
                result = process_attachment(
                    attachment_data=image_bytes,
                    attachment_name=image_filename,
                    alt_text=alt_text,
                    attachment_mode=options.attachment_mode,
                    attachment_output_dir=options.attachment_output_dir,
                    attachment_base_url=options.attachment_base_url,
                    is_image=True,
                    alt_text_mode=options.alt_text_mode,
                )

                # Collect footnote info if present
                if result.get("footnote_label") and result.get("footnote_content"):
                    collected_footnotes[result["footnote_label"]] = result["footnote_content"]

                # Create Image AST node using helper
                image_node = attachment_result_to_image_node(result, fallback_alt_text=alt_text)
                if image_node and isinstance(image_node, Image):
                    images.append(image_node)

            except Exception as e:
                logger.debug(f"Failed to extract image from XLSX: {e!r}")
                continue

    except Exception as e:
        logger.debug(f"Error accessing sheet images: {e!r}")

    return images, collected_footnotes


def _extract_sheet_charts(sheet: Any, base_filename: str, options: Any) -> list[Node]:
    """Extract charts from an XLSX sheet and convert to AST nodes.

    Parameters
    ----------
    sheet : Any
        Openpyxl worksheet
    base_filename : str
        Base filename for context
    options : XlsxOptions
        Conversion options

    Returns
    -------
    list[Node]
        List of AST nodes (Tables or Paragraphs) representing charts

    """
    chart_nodes: list[Node] = []

    if options.chart_mode == "skip":
        return chart_nodes

    try:
        # NOTE: Using private API (sheet._charts) because openpyxl does not provide
        # a public API for reading existing charts from worksheets. This is the
        # standard approach per openpyxl documentation and community practice.
        # Compatible with openpyxl>=3.1.5. May need updates if openpyxl changes internals.
        charts = []
        if hasattr(sheet, "_charts") and sheet._charts:
            charts = sheet._charts
        elif hasattr(sheet, "charts") and sheet.charts:
            charts = sheet.charts

        if not charts:
            return chart_nodes

        for chart in charts:
            try:
                if options.chart_mode == "data":
                    # Extract chart data and convert to table
                    # Note: Chart object structure uses internal openpyxl APIs
                    # (series, values, numRef, etc.) which may vary by version
                    table_node = _chart_to_table_ast(chart)
                    if table_node:
                        chart_nodes.append(table_node)

            except Exception as e:
                logger.debug(f"Failed to extract chart from XLSX: {e!r}")
                continue

    except Exception as e:
        logger.debug(f"Error accessing sheet charts: {e!r}")

    return chart_nodes


def _chart_to_table_ast(chart: Any) -> Table | None:
    """Convert an XLSX chart to a Table AST node.

    Parameters
    ----------
    chart : Any
        Openpyxl chart object

    Returns
    -------
    Table | None
        Table AST node representing chart data, or None if extraction fails

    """
    try:
        # Extract series data
        series_data: list[tuple[str, list[Any]]] = []
        categories: list[str] = []

        if hasattr(chart, "series") and chart.series:
            for series in chart.series:
                series_name = getattr(series, "title", "") or f"Series {len(series_data) + 1}"
                if hasattr(series_name, "v"):
                    series_name = str(series_name.v)

                # Extract values
                values = []
                if hasattr(series, "values") and series.values:
                    values_ref = series.values
                    if hasattr(values_ref, "numRef") and values_ref.numRef:
                        num_cache = values_ref.numRef.numCache
                        if num_cache and hasattr(num_cache, "pt"):
                            values = [pt.v for pt in num_cache.pt if hasattr(pt, "v")]

                series_data.append((str(series_name), values))

                # Extract categories from first series
                if not categories and hasattr(series, "cat") and series.cat:
                    cat_ref = series.cat
                    if hasattr(cat_ref, "strRef") and cat_ref.strRef:
                        str_cache = cat_ref.strRef.strCache
                        if str_cache and hasattr(str_cache, "pt"):
                            categories = [str(pt.v) for pt in str_cache.pt if hasattr(pt, "v")]

        if not series_data:
            return None

        # Use helper to build table from chart data
        return build_chart_table(categories, series_data, category_header="Category")

    except Exception as e:
        logger.debug(f"Failed to convert chart to table: {e!r}")
        return None


class XlsxToAstConverter(BaseParser):
    """Convert XLSX Excel files to AST representation.

    This converter handles XLSX (Excel) format by building Table nodes
    from spreadsheet data.

    Parameters
    ----------
    options : XlsxOptions or None
        Conversion options

    """

    def __init__(self, options: Optional[XlsxOptions] = None, progress_callback: Optional[ProgressCallback] = None):
        """Initialize the XLSX parser with options and progress callback."""
        BaseParser._validate_options_type(options, XlsxOptions, "xlsx")
        options = options or XlsxOptions()
        super().__init__(options, progress_callback)
        self._attachment_footnotes: dict[str, str] = {}  # label -> content for footnote definitions

        # Type hint for IDE
        self.options: XlsxOptions = options

    @requires_dependencies("xlsx", DEPS_XLSX)
    def parse(self, input_data: Union[str, Path, IO[bytes], bytes]) -> Document:
        """Parse the XLSX file into an AST.

        Parameters
        ----------
        input_data : str, Path, IO[bytes], or bytes
            The input Excel file to parse. Can be:
            - File path (str or Path)
            - File-like object in binary mode
            - Raw document bytes

        Returns
        -------
        Document
            AST Document node representing the parsed spreadsheet structure

        Raises
        ------
        ParsingError
            If parsing fails due to invalid format or corruption
        DependencyError
            If required dependencies are not installed
        ValidationError
            If input data is invalid or inaccessible

        """
        import openpyxl

        # Validate ZIP archive security for all input types
        self._validate_zip_security(input_data, suffix=".xlsx")

        # Load workbook
        try:
            doc_input, input_type = validate_and_convert_input(
                input_data, supported_types=["path-like", "file-like", "bytes"], require_binary=True
            )
            wb = openpyxl.load_workbook(doc_input, data_only=self.options.render_formulas)
        except Exception as e:
            raise MalformedFileError(f"Failed to parse XLSX file: {e!r}", original_error=e) from e

        return self.xlsx_to_ast(wb)

    def xlsx_to_ast(self, workbook: Any) -> Document:
        """Convert an openpyxl workbook to AST Document.

        Parameters
        ----------
        workbook : Any
            Openpyxl workbook

        Returns
        -------
        Document
            AST document with table nodes

        """
        children: list[Node] = []

        # Extract metadata
        metadata = self.extract_metadata(workbook)

        # Determine base filename for attachments
        base_filename = "spreadsheet"
        try:
            if hasattr(workbook, "properties") and workbook.properties and workbook.properties.title:
                base_filename = workbook.properties.title.replace(" ", "_")
        except Exception:
            pass

        # Create attachment sequencer for unique filenames
        attachment_sequencer = create_attachment_sequencer()

        # Select sheets
        sheet_names: list[str] = list(workbook.sheetnames)
        if isinstance(self.options.sheets, list):
            sheet_names = [n for n in sheet_names if n in self.options.sheets]
        elif isinstance(self.options.sheets, str):
            pattern = re.compile(self.options.sheets)
            sheet_names = [n for n in sheet_names if pattern.search(n)]

        if not sheet_names:
            return Document(children=[], metadata=metadata.to_dict())

        for sname in sheet_names:
            sheet = workbook[sname]

            # Add sheet title if requested
            if self.options.include_sheet_titles:
                # Create heading node
                children.append(Heading(level=2, content=[Text(content=sname)]))

            # Process sheet data
            # Handle merged cells based on mode
            merged_map: dict[str, str] = {}
            if self.options.merged_cell_mode != "skip":
                merged_map = _map_merged_cells(sheet)

            raw_rows: list[list[Any]] = []
            for row in _xlsx_iter_rows(sheet, self.options.max_rows, self.options.max_cols):
                if all((cell.value is None) for cell in row):
                    raw_rows.append(row)
                else:
                    raw_rows.append(row)

            if not raw_rows:
                continue

            # Convert to strings and handle merged cells based on mode
            str_rows: list[list[str]] = []
            for row in raw_rows:
                out: list[str] = []
                for cell in row:
                    coord = getattr(cell, "coordinate", None)
                    # Only apply merged cell logic if mode is "flatten"
                    if (
                        self.options.merged_cell_mode == "flatten"
                        and coord
                        and coord in merged_map
                        and merged_map[coord] != coord
                    ):
                        out.append("")
                    else:
                        out.append(_format_link_or_text(cell, cell.value, self.options.preserve_newlines_in_cells))
                str_rows.append(out)

            # Trim empty rows based on trim_empty option
            str_rows = trim_rows(str_rows, cast(Any, self.options.trim_empty))

            if not str_rows:
                continue

            # Trim empty columns based on trim_empty option
            str_rows = trim_columns(str_rows, cast(Any, self.options.trim_empty))

            if not str_rows or not any(str_rows):
                continue

            header = str_rows[0]
            data = str_rows[1:] if len(str_rows) > 1 else []

            # Apply header case transformation
            header = transform_header_case(header, self.options.header_case)

            # Compute alignments from header cells
            alignments: list[Alignment] = []
            try:
                first_row_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
                for c_idx, cell in enumerate(first_row_cells, start=1):
                    if self.options.max_cols is not None and c_idx > self.options.max_cols:
                        break
                    alignments.append(_alignment_for_cell(cell))
            except Exception:
                alignments = cast(list[Alignment], ["center"] * len(header))

            # Build table AST
            table = build_table_ast(header, data, alignments)
            children.append(table)

            # Add truncation indicator as paragraph if needed
            truncated_rows = self.options.max_rows is not None and sheet.max_row > self.options.max_rows
            truncated_cols = self.options.max_cols is not None and sheet.max_column > self.options.max_cols
            if truncated_rows or truncated_cols:
                children.append(Paragraph(content=[HTMLInline(content=f"*{self.options.truncation_indicator}*")]))

            # Extract images from sheet
            sheet_images, sheet_footnotes = _extract_sheet_images(
                sheet, base_filename, attachment_sequencer, self.options
            )
            self._attachment_footnotes.update(sheet_footnotes)
            children.extend(sheet_images)

            # Extract charts from sheet
            sheet_charts = _extract_sheet_charts(sheet, base_filename, self.options)
            children.extend(sheet_charts)

        # Append attachment footnote definitions if any were collected
        if self.options.attachments_footnotes_section:
            self._append_attachment_footnotes(
                children, self._attachment_footnotes, self.options.attachments_footnotes_section
            )

        return Document(children=children, metadata=metadata.to_dict())

    def extract_metadata(self, document: Any) -> DocumentMetadata:
        """Extract metadata from XLSX workbook.

        Parameters
        ----------
        document : Any
            Excel workbook (openpyxl)

        Returns
        -------
        DocumentMetadata
            Extracted metadata

        """
        metadata = DocumentMetadata()

        # Try to extract metadata from workbook properties
        try:
            props = document.properties
            if props:
                if props.title:
                    metadata.title = props.title
                if props.creator:
                    metadata.author = props.creator
                if props.subject:
                    metadata.subject = props.subject
                if props.keywords:
                    metadata.keywords = props.keywords.split(",") if isinstance(props.keywords, str) else []
                if props.created:
                    metadata.creation_date = str(props.created)
        except Exception:
            pass

        # Add custom metadata
        try:
            metadata.custom["sheet_count"] = len(document.sheetnames)
            metadata.custom["sheet_names"] = document.sheetnames
        except Exception:
            pass

        return metadata


# Converter metadata for registration
CONVERTER_METADATA = ConverterMetadata(
    format_name="xlsx",
    extensions=[".xlsx"],
    mime_types=["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"],
    magic_bytes=[
        (b"PK\x03\x04", 0),
    ],
    parser_class=XlsxToAstConverter,
    renderer_class=None,
    parser_required_packages=[("openpyxl", "openpyxl", "")],
    renderer_required_packages=[],
    import_error_message="XLSX conversion requires 'openpyxl'. Install with: pip install openpyxl",
    parser_options_class=XlsxOptions,
    renderer_options_class=None,
    description="Convert Excel XLSX files to Markdown tables",
    priority=6,
)
