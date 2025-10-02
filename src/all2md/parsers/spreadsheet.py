#  Copyright (c) 2025 Tom Villani, Ph.D.
#
# src/all2md/parsers/spreadsheet.py
"""Spreadsheet to AST converter.

This module provides conversion from spreadsheet formats (XLSX, ODS, CSV, TSV)
to AST representation. It replaces direct markdown string generation with
structured AST building, enabling multiple rendering strategies.

"""

from __future__ import annotations

import csv
import io
import logging
import re
from pathlib import Path
from typing import IO, Any, Iterable, Optional, Union

from all2md import DependencyError, InputError
from all2md.ast import Document, Emphasis, Heading, HTMLInline, Paragraph, Table, TableCell, TableRow, Text
from all2md.constants import TABLE_ALIGNMENT_MAPPING
from all2md.converter_metadata import ConverterMetadata
from all2md.options import MarkdownOptions, SpreadsheetOptions
from all2md.parsers.base import BaseParser
from all2md.utils.inputs import validate_and_convert_input
from all2md.utils.metadata import DocumentMetadata

logger = logging.getLogger(__name__)


def _sanitize_cell_text(text: Any, preserve_newlines: bool = False) -> str:
    """Convert any cell value to a safe string for AST Text node.

    Note: Markdown escaping is handled by the renderer, not here.
    We only normalize whitespace and convert to string.

    Parameters
    ----------
    text : Any
        Cell value to sanitize
    preserve_newlines : bool
        If True, preserve newlines as <br> tags; if False, replace with spaces

    Returns
    -------
    str
        Sanitized cell text

    """
    if text is None:
        s = ""
    else:
        s = str(text)

    # Handle newlines based on preserve_newlines option
    if preserve_newlines:
        # Keep line breaks as <br> tags
        s = s.replace("\r\n", "<br>").replace("\r", "<br>").replace("\n", "<br>")
    else:
        # Normalize whitespace/newlines inside cells
        s = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")

    return s


def _format_link_or_text(cell: Any, text: str, preserve_newlines: bool = False) -> str:
    """Get cell text, checking for hyperlinks.

    Parameters
    ----------
    cell : Any
        Cell object (openpyxl)
    text : str
        Cell text value
    preserve_newlines : bool
        Whether to preserve newlines in cells

    Returns
    -------
    str
        Cell text (hyperlinks not yet supported in AST path)

    """
    # TODO: Support hyperlinks in table cells via Link nodes
    # For now, just return sanitized text
    return _sanitize_cell_text(text, preserve_newlines)


def _alignment_for_cell(cell: Any) -> str:
    """Map an openpyxl cell alignment to alignment string.

    Parameters
    ----------
    cell : Any
        Cell object

    Returns
    -------
    str
        Alignment: 'left', 'center', 'right', or 'center' (default)

    """
    try:
        align = getattr(cell, "alignment", None)
        if align and getattr(align, "horizontal", None):
            horiz = align.horizontal.lower()
            # Map Excel alignment to our alignment values
            alignment_map = {
                "left": "left",
                "center": "center",
                "centre": "center",
                "right": "right",
                "general": "left",
                "justify": "left",
            }
            return alignment_map.get(horiz, "center")
    except Exception:
        pass

    return "center"


def _build_table_ast(
    header: list[str], rows: list[list[str]], alignments: list[str]
) -> Table:
    """Build an AST Table from header, rows, and alignments.

    Parameters
    ----------
    header : list[str]
        Header row cells
    rows : list[list[str]]
        Data rows
    alignments : list[str]
        Column alignments ('left', 'center', 'right')

    Returns
    -------
    Table
        AST Table node

    """
    # Build header row
    header_cells = [
        TableCell(content=[Text(content=cell)], alignment=alignments[i] if i < len(alignments) else "center")
        for i, cell in enumerate(header)
    ]
    header_row = TableRow(cells=header_cells, is_header=True)

    # Build data rows
    data_rows = []
    for row in rows:
        row_cells = [
            TableCell(content=[Text(content=cell)], alignment=alignments[i] if i < len(alignments) else "center")
            for i, cell in enumerate(row)
        ]
        data_rows.append(TableRow(cells=row_cells, is_header=False))

    # Convert alignment strings to the format expected by Table node
    # Table expects Optional[Literal['left', 'center', 'right']]
    table_alignments: list[Optional[str]] = []
    for align in alignments:
        if align in ("left", "center", "right"):
            table_alignments.append(align)
        else:
            table_alignments.append("center")

    return Table(header=header_row, rows=data_rows, alignments=table_alignments)


def _xlsx_iter_rows(sheet: Any, max_rows: int | None, max_cols: int | None) -> Iterable[list[Any]]:
    """Iterate rows of an openpyxl worksheet with optional truncation.

    Parameters
    ----------
    sheet : Any
        Openpyxl worksheet
    max_rows : int | None
        Maximum rows to read
    max_cols : int | None
        Maximum columns to read

    Yields
    ------
    list[Any]
        Row cells

    """
    row_iter = sheet.iter_rows(values_only=False)
    for r_idx, row in enumerate(row_iter, start=1):
        if max_rows is not None and r_idx > max_rows:
            break
        out_row = []
        for c_idx, cell in enumerate(row, start=1):
            if max_cols is not None and c_idx > max_cols:
                break
            out_row.append(cell)
        yield out_row


def _map_merged_cells(sheet: Any) -> dict[str, str]:
    """Return a map of cell.coordinate -> master.coordinate for merged ranges.

    Parameters
    ----------
    sheet : Any
        Openpyxl worksheet

    Returns
    -------
    dict[str, str]
        Mapping of cell coordinates to master cell coordinates

    """
    merged_map: dict[str, str] = {}
    try:
        ranges = getattr(sheet, "merged_cells", None)
        if not ranges or not getattr(ranges, "ranges", None):
            return merged_map

        for mcr in ranges.ranges:
            min_r, min_c, max_r, max_c = mcr.min_row, mcr.min_col, mcr.max_row, mcr.max_col
            master = sheet.cell(row=min_r, column=min_c).coordinate
            for rr in range(min_r, max_r + 1):
                for cc in range(min_c, max_c + 1):
                    coord = sheet.cell(row=rr, column=cc).coordinate
                    merged_map[coord] = master
    except Exception as e:
        logger.debug(f"Unable to compute merged cell map: {e!r}")
    return merged_map


class SpreadsheetToAstConverter(BaseParser):
    """Convert spreadsheet formats to AST representation.

    This converter handles XLSX, ODS, CSV, and TSV formats by building
    Table nodes from spreadsheet data.

    Parameters
    ----------
    options : SpreadsheetOptions or None
        Conversion options

    """

    def __init__(self, options: SpreadsheetOptions | None = None):
        options = options or SpreadsheetOptions()
        super().__init__()
        self.options: SpreadsheetOptions = options

    def parse(self, input_data: Union[str, Path, IO[bytes], bytes]) -> Document:
        """Parse the input spreadsheet into an AST.

        This method detects the spreadsheet format (XLSX, ODS, CSV, TSV) and
        routes to the appropriate parsing method.

        Parameters
        ----------
        input_data : str, Path, IO[bytes], or bytes
            The input spreadsheet to parse. Can be:
            - File path (str or Path)
            - File-like object in binary mode
            - Raw document bytes

        Returns
        -------
        Document
            AST Document node representing the parsed spreadsheet structure

        Raises
        ------
        MarkdownConversionError
            If parsing fails due to invalid format or corruption
        DependencyError
            If required dependencies are not installed
        InputError
            If input data is invalid or inaccessible

        """

        # Detect format
        detected_format = self._detect_format(input_data)

        try:
            if detected_format == "xlsx":
                # Import openpyxl
                try:
                    import openpyxl
                except ImportError as e:
                    raise DependencyError(
                        converter_name="xlsx",
                        missing_packages=[("openpyxl", "")]
                    ) from e

                # Load workbook
                doc_input, input_type = validate_and_convert_input(
                    input_data, supported_types=["path-like", "file-like", "bytes"], require_binary=True
                )
                wb = openpyxl.load_workbook(doc_input, data_only=self.options.render_formulas)
                return self.xlsx_to_ast(wb)

            elif detected_format == "ods":
                # Import odfpy
                try:
                    from odf import opendocument
                except ImportError as e:
                    raise DependencyError(
                        converter_name="ods",
                        missing_packages=[("odfpy", "")]
                    ) from e

                # Load ODS document
                doc_input, input_type = validate_and_convert_input(
                    input_data, supported_types=["path-like", "file-like", "bytes"], require_binary=True
                )
                doc = opendocument.load(doc_input)
                return self.ods_to_ast(doc)

            elif detected_format == "csv":
                return self.csv_or_tsv_to_ast(input_data, delimiter=",", force_delimiter=False)

            elif detected_format == "tsv":
                force = True
                if self.options.detect_csv_dialect:
                    force = False
                return self.csv_or_tsv_to_ast(input_data, delimiter="\t", force_delimiter=force)

            else:
                # Fallback to CSV
                return self.csv_or_tsv_to_ast(input_data, delimiter=",", force_delimiter=False)

        except (DependencyError, InputError):
            raise
        except Exception as e:
            raise InputError(
                f"Failed to parse spreadsheet: {e!r}",
                original_error=e
            ) from e

    def _detect_format(self, input_data: Union[str, Path, IO[bytes], bytes]) -> str:
        """Detect spreadsheet format from input data.

        Parameters
        ----------
        input_data : Union[str, Path, IO[bytes], bytes]
            Input data to analyze

        Returns
        -------
        str
            Format name: "xlsx", "ods", "csv", or "tsv"

        """
        # If it's a path, check extension first
        if isinstance(input_data, (str, Path)):
            path = Path(input_data)
            ext = path.suffix.lower()
            if ext == ".xlsx":
                return "xlsx"
            elif ext == ".ods":
                return "ods"
            elif ext == ".csv":
                return "csv"
            elif ext == ".tsv":
                return "tsv"

        # Check if file object has a name attribute
        elif hasattr(input_data, 'name'):
            filename = getattr(input_data, 'name', None)
            if filename:
                path = Path(filename)
                ext = path.suffix.lower()
                if ext == ".xlsx":
                    return "xlsx"
                elif ext == ".ods":
                    return "ods"
                elif ext == ".csv":
                    return "csv"
                elif ext == ".tsv":
                    return "tsv"

        # Try content-based detection
        try:
            sample: bytes = b""
            if hasattr(input_data, 'read') and not isinstance(input_data, (str, Path)):
                pos = getattr(input_data, 'tell', lambda: 0)()
                try:
                    input_data.seek(0)
                    sample_data = input_data.read(1024)
                    input_data.seek(pos)
                    if isinstance(sample_data, bytes):
                        sample = sample_data
                    elif isinstance(sample_data, str):
                        sample = sample_data.encode('utf-8', errors='ignore')
                except Exception:
                    pass
            elif isinstance(input_data, bytes):
                sample = input_data[:1024]
            elif isinstance(input_data, (str, Path)):
                try:
                    with open(str(input_data), 'rb') as f:
                        sample = f.read(1024)
                except Exception:
                    pass

            # Check for ZIP-based formats (XLSX or ODS)
            if sample.startswith(b'PK\x03\x04'):
                return "xlsx"  # Default to XLSX for ZIP files

            # Check for CSV/TSV patterns
            try:
                text_sample = sample.decode('utf-8', errors='ignore')
                lines = text_sample.split('\n')[:5]
                non_empty_lines = [line for line in lines if line.strip()]

                if len(non_empty_lines) >= 2:
                    comma_count = sum(line.count(',') for line in non_empty_lines)
                    tab_count = sum(line.count('\t') for line in non_empty_lines)

                    if tab_count >= len(non_empty_lines):
                        return "tsv"
                    elif comma_count >= len(non_empty_lines):
                        return "csv"
            except UnicodeDecodeError:
                pass
        except Exception:
            pass

        # Default fallback to CSV
        return "csv"

    def xlsx_to_ast(self, workbook: Any) -> Document:
        """Convert an openpyxl workbook to AST Document.

        Parameters
        ----------
        workbook : Any
            Openpyxl workbook

        Returns
        -------
        Document
            AST document with table nodes

        """
        children = []

        # Extract metadata
        metadata = self.extract_metadata(workbook)

        # Select sheets
        sheet_names: list[str] = list(workbook.sheetnames)
        if isinstance(self.options.sheets, list):
            sheet_names = [n for n in sheet_names if n in self.options.sheets]
        elif isinstance(self.options.sheets, str):
            pattern = re.compile(self.options.sheets)
            sheet_names = [n for n in sheet_names if pattern.search(n)]

        if not sheet_names:
            return Document(children=[], metadata=metadata.to_dict())

        for sname in sheet_names:
            sheet = workbook[sname]

            # Add sheet title if requested
            if self.options.include_sheet_titles:
                # Create heading node
                children.append(Heading(level=2, content=[Text(content=sname)]))

            # Process sheet data
            merged_map = _map_merged_cells(sheet)

            raw_rows: list[list[Any]] = []
            for row in _xlsx_iter_rows(sheet, self.options.max_rows, self.options.max_cols):
                if all((cell.value is None) for cell in row):
                    raw_rows.append(row)
                else:
                    raw_rows.append(row)

            if not raw_rows:
                continue

            # Convert to strings and handle merged cells
            str_rows: list[list[str]] = []
            for row in raw_rows:
                out: list[str] = []
                for cell in row:
                    coord = getattr(cell, "coordinate", None)
                    if coord and coord in merged_map and merged_map[coord] != coord:
                        out.append("")
                    else:
                        out.append(_format_link_or_text(cell, cell.value, self.options.preserve_newlines_in_cells))
                str_rows.append(out)

            # Trim empty rows based on trim_empty option
            str_rows = self._trim_rows(str_rows)

            if not str_rows:
                continue

            # Trim empty columns based on trim_empty option
            str_rows = self._trim_columns(str_rows)

            if not str_rows or not any(str_rows):
                continue

            header = str_rows[0]
            data = str_rows[1:] if len(str_rows) > 1 else []

            # Apply header case transformation
            header = self._transform_header_case(header)

            # Compute alignments from header cells
            alignments: list[str] = []
            try:
                first_row_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
                for c_idx, cell in enumerate(first_row_cells, start=1):
                    if self.options.max_cols is not None and c_idx > self.options.max_cols:
                        break
                    alignments.append(_alignment_for_cell(cell))
            except Exception:
                alignments = ["center"] * len(header)

            # Build table AST
            table = _build_table_ast(header, data, alignments)
            children.append(table)

            # Add truncation indicator as paragraph if needed
            truncated_rows = self.options.max_rows is not None and sheet.max_row > self.options.max_rows
            truncated_cols = self.options.max_cols is not None and sheet.max_column > self.options.max_cols
            if truncated_rows or truncated_cols:
                children.append(
                    Paragraph(content=[HTMLInline(content=f"*{self.options.truncation_indicator}*")])
                )

        return Document(children=children, metadata=metadata.to_dict())

    def csv_or_tsv_to_ast(
        self,
        input_data: Union[str, Path, IO[bytes], IO[str]],
        delimiter: str | None = None,
        force_delimiter: bool = False,
    ) -> Document:
        """Convert CSV/TSV to AST Document.

        Parameters
        ----------
        input_data : Union[str, Path, IO[bytes], IO[str]]
            Input data
        delimiter : str | None
            Delimiter character to use (e.g., ',' for CSV, '\\t' for TSV)
        force_delimiter : bool
            If True, forces use of the specified delimiter without dialect detection.
            For TSV files, this defaults to True (forces tab delimiter) unless
            options.detect_csv_dialect is explicitly enabled.

        Returns
        -------
        Document
            AST document with table node

        Notes
        -----
        Delimiter handling behavior:
        - If options.csv_delimiter is set: Uses that delimiter regardless
        - If force_delimiter=True: Uses specified delimiter without detection
        - If options.detect_csv_dialect=True: Attempts automatic dialect detection
        - Otherwise: Uses specified delimiter or defaults to Excel dialect

        For TSV files called via tsv_to_markdown(), force_delimiter is True by
        default to ensure tab (\\t) delimiter is used, unless detect_csv_dialect
        is explicitly enabled in options.

        """
        from all2md.utils.inputs import validate_and_convert_input

        # Validate and load text
        try:
            doc_input, _ = validate_and_convert_input(input_data, supported_types=["path-like", "file-like", "bytes"])
            if isinstance(doc_input, (str, Path)):
                with open(doc_input, "rb") as f:
                    text_stream = self._read_text_stream_for_csv(f)
            else:
                text_stream = self._read_text_stream_for_csv(doc_input)
        except Exception as e:
            from all2md.exceptions import MarkdownConversionError
            raise MarkdownConversionError(
                f"Failed to read CSV/TSV input: {e}", conversion_stage="input_processing", original_error=e
            ) from e

        # Sniff dialect
        text_stream.seek(0)
        sample = text_stream.read(4096)
        text_stream.seek(0)

        dialect: csv.Dialect
        if self.options.csv_delimiter:
            exc_dialect = csv.excel()
            exc_dialect.delimiter = self.options.csv_delimiter
            dialect = exc_dialect
        elif force_delimiter:
            exc_dialect = csv.excel()
            exc_dialect.delimiter = delimiter or "\t"
            dialect = exc_dialect
        elif self.options.detect_csv_dialect:
            try:
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(sample, delimiters=",\t;|\x1f")
            except Exception:
                exc_dialect = csv.excel()
                if delimiter:
                    exc_dialect.delimiter = delimiter
                dialect = exc_dialect
        else:
            exc_dialect = csv.excel()
            if delimiter:
                exc_dialect.delimiter = delimiter
            dialect = exc_dialect

        reader = csv.reader(text_stream, dialect=dialect)
        rows: list[list[str]] = []
        for r in reader:
            rows.append(r)

        # Extract metadata (CSV/TSV have no structured metadata)
        metadata = self.extract_metadata(None)

        # Drop leading/trailing fully empty rows
        def is_empty(row: list[str]) -> bool:
            return all((not (c or "").strip()) for c in row)

        while rows and is_empty(rows[0]):
            rows.pop(0)
        while rows and is_empty(rows[-1]):
            rows.pop()

        if not rows:
            return Document(children=[], metadata=metadata.to_dict())

        # Handle header
        if self.options.has_header:
            header = rows[0]
            data_rows = rows[1:]
        else:
            if rows:
                num_cols = len(rows[0]) if rows else 0
                header = [f"Column {i+1}" for i in range(num_cols)]
                data_rows = rows
            else:
                return Document(children=[], metadata=metadata.to_dict())

        # Truncate columns
        if self.options.max_cols is not None:
            header = header[: self.options.max_cols]
            data_rows = [r[: self.options.max_cols] for r in data_rows]

        # Truncate rows
        truncated = False
        if self.options.max_rows is not None and len(data_rows) > self.options.max_rows:
            data_rows = data_rows[: self.options.max_rows]
            truncated = True

        # Sanitize cells
        if self.options.has_header:
            header = [_sanitize_cell_text(c).lstrip("\ufeff") for c in header]
        else:
            header = [_sanitize_cell_text(c) for c in header]
        data_rows = [[_sanitize_cell_text(c) for c in r] for r in data_rows]

        # Alignments default to center
        alignments = ["center"] * len(header)

        # Build table AST
        table = _build_table_ast(header, data_rows, alignments)

        children = [table]

        if truncated:
            children.append(
                Paragraph(content=[HTMLInline(content=f"*{self.options.truncation_indicator}*")])
            )

        return Document(children=children, metadata=metadata.to_dict())

    def _read_text_stream_for_csv(self, input_data: Any) -> io.StringIO:
        """Read binary or text input and return a StringIO for CSV parsing.

        Parameters
        ----------
        input_data : Any
            Input stream or data

        Returns
        -------
        io.StringIO
            Text stream for CSV parsing

        """
        if isinstance(input_data, io.StringIO):
            input_data.seek(0)
            return input_data

        content = None
        if hasattr(input_data, "read"):
            pos = None
            try:
                pos = input_data.tell()
            except Exception:
                pass

            raw = input_data.read()
            if isinstance(raw, bytes):
                try:
                    content = raw.decode("utf-8-sig")
                except UnicodeDecodeError:
                    content = raw.decode("utf-8", errors="replace")
            else:
                content = str(raw)

            try:
                if pos is not None:
                    input_data.seek(pos)
            except Exception:
                pass
        else:
            content = str(input_data)

        return io.StringIO(content or "")

    def ods_to_ast(self, doc: Any) -> Document:
        """Convert ODS document to AST Document.

        Parameters
        ----------
        doc : Any
            ODS document object

        Returns
        -------
        Document
            AST document with table nodes

        """
        from odf.table import Table as OdfTable, TableCell, TableRow

        children = []

        # Extract metadata
        metadata = self.extract_metadata(doc)

        body = doc.body
        tables = list(body.getElementsByType(OdfTable)) if body else []

        if not tables:
            return Document(children=[], metadata=metadata.to_dict())

        # Filter sheets based on options
        sheet_names = [table.getAttribute("name") or f"Sheet{i+1}" for i, table in enumerate(tables)]
        selected_tables = []
        selected_names = []

        if isinstance(self.options.sheets, list):
            for i, name in enumerate(sheet_names):
                if name in self.options.sheets:
                    selected_tables.append(tables[i])
                    selected_names.append(name)
        elif isinstance(self.options.sheets, str):
            pattern = re.compile(self.options.sheets)
            for i, name in enumerate(sheet_names):
                if pattern.search(name):
                    selected_tables.append(tables[i])
                    selected_names.append(name)
        else:
            selected_tables = tables
            selected_names = sheet_names

        # Process each selected sheet
        for table, name in zip(selected_tables, selected_names, strict=False):
            # Add sheet title if requested
            if self.options.include_sheet_titles:
                children.append(Heading(level=2, content=[Text(content=name)]))

            # Extract rows
            rows_elem = list(table.getElementsByType(TableRow))
            if not rows_elem:
                continue

            # Convert to cell data
            raw_rows = []
            for row in rows_elem:
                cells = list(row.getElementsByType(TableCell))
                row_data = []

                for cell in cells:
                    cell_text = ""
                    for node in cell.childNodes:
                        if hasattr(node, 'data'):
                            cell_text += str(node.data)
                        elif hasattr(node, 'childNodes'):
                            for subnode in node.childNodes:
                                if hasattr(subnode, 'data'):
                                    cell_text += str(subnode.data)

                    # Handle cell repetition
                    repeat_count = 1
                    try:
                        repeat_attr = cell.getAttribute("numbercolumnsrepeated")
                        if repeat_attr:
                            repeat_count = int(repeat_attr)
                    except (ValueError, TypeError):
                        pass

                    for _ in range(repeat_count):
                        row_data.append(cell_text.strip() if cell_text else "")

                raw_rows.append(row_data)

            # Remove empty trailing rows
            while raw_rows and all(not cell for cell in raw_rows[-1]):
                raw_rows.pop()

            if not raw_rows:
                continue

            # Apply row limits
            if self.options.max_rows is not None:
                total_available = len(raw_rows) - 1
                if total_available > self.options.max_rows:
                    raw_rows = raw_rows[: self.options.max_rows + 1]

            # Extract header and data
            if self.options.has_header:
                header = raw_rows[0]
                data_rows = raw_rows[1:] if len(raw_rows) > 1 else []
            else:
                if raw_rows:
                    max_cols = max(len(row) for row in raw_rows)
                    if self.options.max_cols is not None:
                        max_cols = min(max_cols, self.options.max_cols)
                    header = [f"Column {i+1}" for i in range(max_cols)]
                    data_rows = raw_rows
                else:
                    header = []
                    data_rows = []

            # Apply column limits
            if self.options.max_cols is not None:
                header = header[: self.options.max_cols]
                data_rows = [row[: self.options.max_cols] for row in data_rows]

            # Apply row/column trimming
            all_rows = [header] + data_rows if header else data_rows
            all_rows = self._trim_rows(all_rows)
            all_rows = self._trim_columns(all_rows)

            if not all_rows:
                continue

            # Sanitize all cell content
            header = [_sanitize_cell_text(cell, self.options.preserve_newlines_in_cells) for cell in all_rows[0]]
            data_rows = [
                [_sanitize_cell_text(cell, self.options.preserve_newlines_in_cells) for cell in row]
                 for row in all_rows[1:]
            ]

            # Apply header case transformation
            header = self._transform_header_case(header)

            # Ensure all rows have same number of columns
            if header:
                max_cols = len(header)
                data_rows = [row + [""] * (max_cols - len(row)) for row in data_rows]

            # Build table
            if header:
                alignments = ["center"] * len(header)
                table_node = _build_table_ast(header, data_rows, alignments)
                children.append(table_node)

            # Add truncation indicator if needed
            truncated = (self.options.max_rows is not None and len(rows_elem) - 1 > self.options.max_rows) or \
                       (self.options.max_cols is not None and any(len(row) > self.options.max_cols for row in raw_rows))
            if truncated:
                children.append(
                    Paragraph(content=[HTMLInline(content=f"*{self.options.truncation_indicator}*")])
                )

        return Document(children=children, metadata=metadata.to_dict())

    def _trim_rows(self, rows: list[list[str]]) -> list[list[str]]:
        """Trim empty rows based on trim_empty option.

        Parameters
        ----------
        rows : list[list[str]]
            Rows to trim

        Returns
        -------
        list[list[str]]
            Trimmed rows

        """
        if not rows or self.options.trim_empty == "none":
            return rows

        # Trim leading empty rows
        if self.options.trim_empty in ("leading", "both"):
            while rows and all(c == "" for c in rows[0]):
                rows.pop(0)

        # Trim trailing empty rows
        if self.options.trim_empty in ("trailing", "both"):
            while rows and all(c == "" for c in rows[-1]):
                rows.pop()

        return rows

    def _trim_columns(self, rows: list[list[str]]) -> list[list[str]]:
        """Trim empty columns based on trim_empty option.

        Parameters
        ----------
        rows : list[list[str]]
            Rows to trim columns from

        Returns
        -------
        list[list[str]]
            Rows with trimmed columns

        """
        if not rows or self.options.trim_empty == "none":
            return rows

        if not rows[0]:
            return rows

        num_cols = len(rows[0])

        # Find leading empty columns
        leading_empty = 0
        if self.options.trim_empty in ("leading", "both"):
            for col_idx in range(num_cols):
                if all(row[col_idx] == "" if col_idx < len(row) else True for row in rows):
                    leading_empty += 1
                else:
                    break

        # Find trailing empty columns
        trailing_empty = 0
        if self.options.trim_empty in ("trailing", "both"):
            for col_idx in range(num_cols - 1, -1, -1):
                if all(row[col_idx] == "" if col_idx < len(row) else True for row in rows):
                    trailing_empty += 1
                else:
                    break

        # Trim columns
        if leading_empty > 0 or trailing_empty > 0:
            end_col = num_cols - trailing_empty
            return [row[leading_empty:end_col] for row in rows]

        return rows

    def _transform_header_case(self, header: list[str]) -> list[str]:
        """Transform header case based on header_case option.

        Parameters
        ----------
        header : list[str]
            Header row

        Returns
        -------
        list[str]
            Transformed header

        """
        if self.options.header_case == "preserve":
            return header
        elif self.options.header_case == "title":
            return [cell.title() for cell in header]
        elif self.options.header_case == "upper":
            return [cell.upper() for cell in header]
        elif self.options.header_case == "lower":
            return [cell.lower() for cell in header]
        return header

    def extract_metadata(self, document: Any) -> DocumentMetadata:
        """Extract metadata from spreadsheet document.

        Parameters
        ----------
        document : Any
            Spreadsheet document (workbook, ODS doc, etc.)

        Returns
        -------
        DocumentMetadata
            Empty metadata (spreadsheets don't have standard metadata in this context)

        Notes
        -----
        While spreadsheet formats like XLSX and ODS can contain metadata
        (author, title, etc.), this is typically accessed during the
        parsing phase. The document object passed here is already processed
        into tables, so metadata extraction is not applicable at this stage.

        For format-specific metadata extraction, see the individual
        parsing methods (xlsx_to_ast, ods_to_ast) which could be extended
        to capture metadata during the initial file loading.

        """
        return DocumentMetadata()


def _detect_csv_tsv_content(content: bytes) -> bool:
    """Content-based detector for CSV/TSV formats.

    This function is used by the registry to detect CSV/TSV files
    based on content patterns when file extensions are not available.

    Parameters
    ----------
    content : bytes
        File content to analyze

    Returns
    -------
    bool
        True if content appears to be CSV or TSV

    """
    try:
        content_str = content.decode('utf-8', errors='ignore')
        non_empty_lines = [line.strip() for line in content_str.split('\n') if line.strip()]

        if len(non_empty_lines) >= 2:
            comma_count = sum(line.count(',') for line in non_empty_lines)
            tab_count = sum(line.count('\t') for line in non_empty_lines)

            if comma_count >= len(non_empty_lines):
                logger.debug(f"CSV pattern detected: {comma_count} commas in {len(non_empty_lines)} lines")
                return True
            elif tab_count >= len(non_empty_lines):
                logger.debug(f"TSV pattern detected: {tab_count} tabs in {len(non_empty_lines)} lines")
                return True
    except UnicodeDecodeError:
        pass

    return False


class SpreadsheetConverterMetadata(ConverterMetadata):
    """Specialized metadata for spreadsheet converter with smart dependency checking."""

    def get_required_packages_for_content(
        self,
        content: Optional[bytes] = None,
        input_data: Optional[Union[str, Path, IO[bytes], bytes]] = None
    ) -> list[tuple[str, str, str]]:
        """Get required packages based on detected spreadsheet format.

        For XLSX files, openpyxl is required. For ODS files, odfpy is required.
        For CSV/TSV files, no additional packages are needed beyond the standard library.

        Parameters
        ----------
        content : bytes, optional
            File content sample (may be partial) to analyze for format detection
        input_data : various types, optional
            Original input data (path, file object, or bytes) for accurate detection

        Returns
        -------
        list[tuple[str, str, str]]
            Required packages for the detected format as (install_name, import_name, version_spec) tuples

        """
        if input_data is not None:
            converter = SpreadsheetToAstConverter()
            detected_format = converter._detect_format(input_data)
            if detected_format == "xlsx":
                return [("openpyxl", "openpyxl", "")]
            elif detected_format == "ods":
                return [("odfpy", "odf", "")]
            elif detected_format in ("csv", "tsv"):
                return []

        if content is None:
            return [("openpyxl", "openpyxl", ""), ("odfpy", "odf", "")]

        if content.startswith(b'PK\x03\x04'):
            MIN_ZIP_SIZE = 100

            if len(content) >= MIN_ZIP_SIZE:
                try:
                    import zipfile

                    with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
                        file_list = zf.namelist()
                        if 'mimetype' in file_list:
                            try:
                                mimetype = zf.read('mimetype').decode('utf-8', errors='ignore').strip()
                                if 'opendocument.spreadsheet' in mimetype:
                                    return [("odfpy", "odf", "")]
                            except Exception:
                                pass
                        if '[Content_Types].xml' in file_list or any(f.startswith('xl/') for f in file_list):
                            return [("openpyxl", "openpyxl", "")]
                        if 'META-INF/manifest.xml' in file_list:
                            return [("odfpy", "odf", "")]
                except Exception:
                    pass

            return [("openpyxl", "openpyxl", ""), ("odfpy", "odf", "")]

        if _detect_csv_tsv_content(content):
            return []

        return [("openpyxl", "openpyxl", "")]


CONVERTER_METADATA = SpreadsheetConverterMetadata(
    format_name="spreadsheet",
    extensions=[".xlsx", ".ods", ".csv", ".tsv"],
    mime_types=[
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.oasis.opendocument.spreadsheet",
        "text/csv",
        "application/csv",
        "text/tab-separated-values"
    ],
    magic_bytes=[
        (b"PK\x03\x04", 0),
    ],
    content_detector=_detect_csv_tsv_content,
    parser_class="SpreadsheetToAstConverter",
    renderer_class=None,
    required_packages=[("openpyxl", "openpyxl", ""), ("odfpy", "odf", "")],
    import_error_message="Spreadsheet conversion requires dependencies: 'openpyxl' for XLSX, "
                        "'odfpy' for ODS. Install with: pip install openpyxl odfpy",
    options_class="SpreadsheetOptions",
    description="Convert spreadsheets (XLSX, ODS, CSV, TSV) to Markdown tables",
    priority=5
)
